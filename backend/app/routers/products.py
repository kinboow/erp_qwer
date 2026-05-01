"""产品列表查询 API — 本地数据库"""

from __future__ import annotations

import io
import logging
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user
from app.models import User
from app.services.erp_sync import ensure_tables

logger = logging.getLogger(__name__)

router = APIRouter(tags=["产品列表"])


@router.get("/", summary="产品列表（分页 + 搜索）")
def api_list_products(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    keyword: Optional[str] = Query(None, description="货号/品名模糊搜索"),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    ensure_tables(db)

    conditions = []
    params: dict[str, Any] = {}

    if keyword:
        conditions.append(
            "(p.product_no LIKE :kw OR p.product_name LIKE :kw OR p.brand LIKE :kw"
            " OR p.product_no IN (SELECT product_no FROM product_name_mappings WHERE alias_name LIKE :kw))"
        )
        params["kw"] = f"%{keyword}%"

    where = " AND ".join(conditions) if conditions else "1=1"

    count_sql = f"SELECT COUNT(*) AS cnt FROM erp_products p WHERE {where}"
    total = db.execute(text(count_sql), params).scalar() or 0

    offset = (page - 1) * page_size
    data_sql = f"""
        SELECT p.id, p.product_id, p.product_no, p.product_name, p.brand, p.category,
               p.color, p.unit, p.price, p.spec, p.material, p.image_url, p.remark, p.is_current_year, p.synced_at,
               (SELECT COUNT(*) FROM product_name_mappings m WHERE m.product_no = p.product_no) AS mapping_count
        FROM erp_products p
        WHERE {where}
        ORDER BY p.product_no ASC, p.id ASC
        LIMIT :limit OFFSET :offset
    """
    params["limit"] = page_size
    params["offset"] = offset
    rows = db.execute(text(data_sql), params).mappings().all()

    products = []
    for r in rows:
        products.append({
            "id": r["id"],
            "product_id": r["product_id"] or "",
            "product_no": r["product_no"] or "",
            "product_name": r["product_name"] or "",
            "brand": r["brand"] or "",
            "category": r["category"] or "",
            "color": r["color"] or "",
            "unit": r["unit"] or "",
            "price": float(r["price"] or 0),
            "spec": r["spec"] or "",
            "material": r["material"] or "",
            "image_url": r["image_url"] or "",
            "remark": r["remark"] or "",
            "is_current_year": bool(r.get("is_current_year")),
            "synced_at": str(r["synced_at"] or ""),
            "mapping_count": r["mapping_count"] or 0,
        })

    return {
        "code": 200,
        "data": {
            "list": products,
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    }


# ---------------------------------------------------------------------------
# 设为本年款 / 取消本年款
# ---------------------------------------------------------------------------

class CurrentYearPayload(BaseModel):
    is_current_year: bool


@router.put("/{product_id}/current-year", summary="设置/取消本年款")
def api_set_current_year(
    product_id: int,
    payload: CurrentYearPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    ensure_tables(db)
    db.execute(
        text("UPDATE erp_products SET is_current_year = :val WHERE id = :id"),
        {"val": 1 if payload.is_current_year else 0, "id": product_id},
    )
    db.commit()
    return {"code": 200, "message": "设置成功"}


@router.post("/batch-current-year", summary="批量设置/取消本年款")
def api_batch_set_current_year(
    payload: dict[str, Any],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    ensure_tables(db)
    ids = payload.get("ids") or []
    is_current_year = payload.get("is_current_year", True)
    if not ids:
        return {"code": 400, "message": "未选择产品"}
    placeholders = ",".join(str(int(i)) for i in ids)
    db.execute(
        text(f"UPDATE erp_products SET is_current_year = :val WHERE id IN ({placeholders})"),
        {"val": 1 if is_current_year else 0},
    )
    db.commit()
    return {"code": 200, "message": f"已更新 {len(ids)} 个产品"}


@router.get("/current-year", summary="本年产品库（分页 + 搜索）")
def api_list_current_year_products(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    keyword: Optional[str] = Query(None, description="货号/品名模糊搜索"),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    ensure_tables(db)

    conditions = ["p.is_current_year = 1"]
    params: dict[str, Any] = {}

    if keyword:
        conditions.append(
            "(p.product_no LIKE :kw OR p.product_name LIKE :kw OR p.brand LIKE :kw)"
        )
        params["kw"] = f"%{keyword}%"

    where = " AND ".join(conditions)

    count_sql = f"SELECT COUNT(*) AS cnt FROM erp_products p WHERE {where}"
    total = db.execute(text(count_sql), params).scalar() or 0

    offset = (page - 1) * page_size
    data_sql = f"""
        SELECT p.id, p.product_id, p.product_no, p.product_name, p.brand, p.category,
               p.color, p.unit, p.price, p.spec, p.material, p.image_url, p.remark, p.is_current_year, p.synced_at,
               (SELECT COUNT(*) FROM product_name_mappings m WHERE m.product_no = p.product_no) AS mapping_count
        FROM erp_products p
        WHERE {where}
        ORDER BY p.product_no ASC, p.id ASC
        LIMIT :limit OFFSET :offset
    """
    params["limit"] = page_size
    params["offset"] = offset
    rows = db.execute(text(data_sql), params).mappings().all()

    products = []
    for r in rows:
        products.append({
            "id": r["id"],
            "product_id": r["product_id"] or "",
            "product_no": r["product_no"] or "",
            "product_name": r["product_name"] or "",
            "brand": r["brand"] or "",
            "category": r["category"] or "",
            "color": r["color"] or "",
            "unit": r["unit"] or "",
            "price": float(r["price"] or 0),
            "spec": r["spec"] or "",
            "material": r["material"] or "",
            "image_url": r["image_url"] or "",
            "remark": r["remark"] or "",
            "is_current_year": bool(r.get("is_current_year")),
            "synced_at": str(r["synced_at"] or ""),
            "mapping_count": r["mapping_count"] or 0,
        })

    return {
        "code": 200,
        "data": {
            "list": products,
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    }


@router.get("/current-year/options", summary="本年产品下拉选项（款号+颜色+尺码）")
def api_current_year_options(
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    """返回去重的款号列表及每个款号对应的颜色和可用尺码列表，用于下拉选择"""
    ensure_tables(db)
    rows = db.execute(text(
        "SELECT product_no, color, spec FROM erp_products WHERE is_current_year = 1 ORDER BY product_no ASC"
    )).mappings().all()
    pno_map: dict[str, dict] = {}
    for r in rows:
        pno = r["product_no"] or ""
        color_raw = r["color"] or ""
        spec_raw = r["spec"] or ""
        if not pno:
            continue
        if pno not in pno_map:
            pno_map[pno] = {"colors": [], "sizes": []}
        for c in color_raw.replace("，", ",").split(","):
            c = c.strip()
            if c and c not in pno_map[pno]["colors"]:
                pno_map[pno]["colors"].append(c)
        for s in spec_raw.replace("，", ",").split(","):
            s = s.strip()
            if s and s not in pno_map[pno]["sizes"]:
                pno_map[pno]["sizes"].append(s)
    product_nos = [{"product_no": k, "colors": v["colors"], "sizes": v["sizes"]} for k, v in pno_map.items()]
    return {"code": 200, "data": product_nos}


# ---------------------------------------------------------------------------
# 名称映射 CRUD
# ---------------------------------------------------------------------------

class AliasPayload(BaseModel):
    alias_name: str


@router.get("/{product_no}/name-mappings", summary="获取货号的名称映射列表")
def api_get_name_mappings(
    product_no: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    ensure_tables(db)
    rows = db.execute(
        text("SELECT id, product_no, alias_name, created_at FROM product_name_mappings WHERE product_no = :pno ORDER BY id ASC"),
        {"pno": product_no},
    ).mappings().all()
    return {
        "code": 200,
        "data": [{"id": r["id"], "product_no": r["product_no"], "alias_name": r["alias_name"], "created_at": str(r["created_at"] or "")} for r in rows],
    }


@router.post("/{product_no}/name-mappings", summary="添加名称映射")
def api_add_name_mapping(
    product_no: str,
    payload: AliasPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    ensure_tables(db)
    alias = payload.alias_name.strip()
    if not alias:
        return {"code": 400, "message": "映射名称不能为空"}

    existing = db.execute(
        text("SELECT id, product_no FROM product_name_mappings WHERE alias_name = :name"),
        {"name": alias},
    ).mappings().first()

    if existing:
        if existing["product_no"] == product_no:
            return {"code": 400, "message": f"该名称已存在于当前货号的映射中"}
        return {"code": 400, "message": f"名称「{alias}」已被货号 {existing['product_no']} 使用，不能重复映射"}

    db.execute(
        text("INSERT INTO product_name_mappings (product_no, alias_name) VALUES (:pno, :name)"),
        {"pno": product_no, "name": alias},
    )
    db.commit()
    return {"code": 200, "message": "添加成功"}


@router.delete("/name-mappings/{mapping_id}", summary="删除名称映射")
def api_delete_name_mapping(
    mapping_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    ensure_tables(db)
    db.execute(text("DELETE FROM product_name_mappings WHERE id = :id"), {"id": mapping_id})
    db.commit()
    return {"code": 200, "message": "删除成功"}


@router.get("/name-mappings/resolve", summary="根据名称解析货号")
def api_resolve_alias(
    name: str = Query(..., description="名称或货号"),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    """先精确匹配货号，再查映射表"""
    ensure_tables(db)
    direct = db.execute(
        text("SELECT product_no FROM erp_products WHERE product_no = :name LIMIT 1"),
        {"name": name},
    ).mappings().first()
    if direct:
        return {"code": 200, "data": {"product_no": direct["product_no"], "matched_by": "product_no"}}

    alias = db.execute(
        text("SELECT product_no FROM product_name_mappings WHERE alias_name = :name LIMIT 1"),
        {"name": name},
    ).mappings().first()
    if alias:
        return {"code": 200, "data": {"product_no": alias["product_no"], "matched_by": "alias"}}

    return {"code": 404, "message": f"未找到名称「{name}」对应的货号"}


# ---------------------------------------------------------------------------
# 导出 Excel
# ---------------------------------------------------------------------------

# ERP 同步字段（只读，导入时忽略修改）
_ERP_READONLY_FIELDS = {"product_id", "brand", "category", "color", "unit", "price", "spec", "material", "image_url", "synced_at"}

# 固定列定义: (excel_header, db_field, is_readonly)
_FIXED_COLUMNS = [
    ("货号", "product_no", True),
    ("品名", "product_name", True),
    ("品牌", "brand", True),
    ("类别", "category", True),
    ("颜色", "color", True),
    ("单位", "unit", True),
    ("单价", "price", True),
    ("尺码", "spec", True),
    ("材质", "material", True),
    ("备注", "remark", True),
    ("本年款", "is_current_year", False),
]


def _get_max_mapping_count(db: Session) -> int:
    """查询所有货号中映射数量的最大值。"""
    row = db.execute(
        text("SELECT MAX(cnt) AS max_cnt FROM (SELECT COUNT(*) AS cnt FROM product_name_mappings GROUP BY product_no) t"),
    ).mappings().first()
    return int(row["max_cnt"] or 0) if row else 0


@router.get("/export", summary="导出产品列表 Excel")
def api_export_products(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ensure_tables(db)

    # 查询所有产品
    rows = db.execute(
        text(
            "SELECT id, product_id, product_no, product_name, brand, category, color, unit, price, spec, material, "
            "image_url, remark, is_current_year, synced_at "
            "FROM erp_products ORDER BY product_no ASC, id ASC"
        ),
    ).mappings().all()

    # 查询所有映射，按 product_no 分组
    mapping_rows = db.execute(
        text("SELECT product_no, alias_name FROM product_name_mappings ORDER BY product_no, id"),
    ).mappings().all()
    mapping_map: dict[str, list[str]] = {}
    for mr in mapping_rows:
        mapping_map.setdefault(mr["product_no"] or "", []).append(mr["alias_name"] or "")

    # 计算映射列数（至少留1列）
    max_mappings = max(_get_max_mapping_count(db), 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品列表"

    # 表头
    headers = [col[0] for col in _FIXED_COLUMNS] + [f"名称映射{i+1}" for i in range(max_mappings)]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    editable_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    editable_font = Font(color="375623", bold=True, size=11)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        is_editable = h == "本年款" or h.startswith("名称映射")
        cell.fill = editable_fill if is_editable else header_fill
        cell.font = editable_font if is_editable else header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, r in enumerate(rows, 2):
        for col_idx, (_, field, _) in enumerate(_FIXED_COLUMNS, 1):
            val = r[field]
            if field == "is_current_year":
                val = "是" if val else ""
            elif field == "price":
                val = float(val or 0)
            else:
                val = str(val or "")
            ws.cell(row=row_idx, column=col_idx, value=val)

        # 映射列
        aliases = mapping_map.get(r["product_no"] or "", [])
        for i in range(max_mappings):
            val = aliases[i] if i < len(aliases) else ""
            ws.cell(row=row_idx, column=len(_FIXED_COLUMNS) + 1 + i, value=val)

    # 列宽自适应
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"},
    )


# ---------------------------------------------------------------------------
# 导入 Excel
# ---------------------------------------------------------------------------

@router.post("/import", summary="导入产品列表 Excel")
async def api_import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    import openpyxl

    ensure_tables(db)

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return {"code": 400, "message": "请上传 .xlsx 格式的 Excel 文件"}

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return {"code": 400, "message": f"Excel 文件解析失败: {exc}"}

    ws = wb.active
    if not ws:
        return {"code": 400, "message": "Excel 中无工作表"}

    # 解析表头
    header_row = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "货号" not in header_row:
        return {"code": 400, "message": "Excel 表头中未找到「货号」列"}

    pno_col = header_row.index("货号")
    current_year_col = header_row.index("本年款") if "本年款" in header_row else None

    # 找出所有「名称映射N」列
    mapping_cols: list[int] = []
    for i, h in enumerate(header_row):
        if h.startswith("名称映射"):
            mapping_cols.append(i)

    updated = 0
    mapping_added = 0
    mapping_removed = 0
    errors: list[str] = []

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        cells = list(row_cells or [])
        if len(cells) <= pno_col:
            continue
        pno = str(cells[pno_col] or "").strip()
        if not pno:
            continue

        # 检查产品是否存在
        product = db.execute(
            text("SELECT id FROM erp_products WHERE product_no = :pno LIMIT 1"),
            {"pno": pno},
        ).mappings().first()
        if not product:
            errors.append(f"第{row_idx}行: 货号「{pno}」不存在，跳过")
            continue

        # 更新本年款
        if current_year_col is not None and len(cells) > current_year_col:
            val = str(cells[current_year_col] or "").strip()
            is_cy = 1 if val in ("是", "1", "true", "True", "yes", "Yes", "TRUE", "YES") else 0
            db.execute(
                text("UPDATE erp_products SET is_current_year = :val WHERE product_no = :pno"),
                {"val": is_cy, "pno": pno},
            )
            updated += 1

        # 处理映射列：以 Excel 为准，同步映射
        if mapping_cols:
            excel_aliases: list[str] = []
            for mc in mapping_cols:
                if mc < len(cells):
                    alias = str(cells[mc] or "").strip()
                    if alias:
                        excel_aliases.append(alias)

            # 查当前数据库中此货号的映射
            existing = db.execute(
                text("SELECT id, alias_name FROM product_name_mappings WHERE product_no = :pno ORDER BY id"),
                {"pno": pno},
            ).mappings().all()
            existing_aliases = {r["alias_name"]: r["id"] for r in existing}

            # 添加新映射
            for alias in excel_aliases:
                if alias not in existing_aliases:
                    # 检查是否被其他货号占用
                    conflict = db.execute(
                        text("SELECT product_no FROM product_name_mappings WHERE alias_name = :name AND product_no != :pno"),
                        {"name": alias, "pno": pno},
                    ).mappings().first()
                    if conflict:
                        errors.append(f"第{row_idx}行: 映射「{alias}」已被货号 {conflict['product_no']} 使用，跳过")
                        continue
                    db.execute(
                        text("INSERT INTO product_name_mappings (product_no, alias_name) VALUES (:pno, :name)"),
                        {"pno": pno, "name": alias},
                    )
                    mapping_added += 1

            # 删除 Excel 中不再存在的映射
            for alias, mid in existing_aliases.items():
                if alias not in excel_aliases:
                    db.execute(text("DELETE FROM product_name_mappings WHERE id = :id"), {"id": mid})
                    mapping_removed += 1

    db.commit()

    summary = f"更新 {updated} 个产品"
    if mapping_added:
        summary += f"，新增 {mapping_added} 个映射"
    if mapping_removed:
        summary += f"，删除 {mapping_removed} 个映射"
    if errors:
        summary += f"，{len(errors)} 条警告"

    return {
        "code": 200,
        "message": summary,
        "data": {
            "updated": updated,
            "mapping_added": mapping_added,
            "mapping_removed": mapping_removed,
            "errors": errors[:50],
        },
    }

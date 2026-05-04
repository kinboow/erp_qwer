import asyncio
import base64
import io
import json
import logging
import secrets
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional

import httpx
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import SessionLocal, run_in_threadpool
from app.models import User, WechatInstance
from app.services.ai_order_parser import AIOrderParserError, ai_order_parser
from app.services.downstream_support import ensure_downstream_support_tables
from app.services.erp_bridge import ERPBridge, ERPBridgeError
from app.services.wechat_event_types import EVENT_TYPE_TO_MESSAGE_TYPE, CONTENT_TYPE_TO_MESSAGE_TYPE

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


erp_bridge = ERPBridge()

REVIEW_UID_PREFIX = "RV"


def _generate_review_uid() -> str:
    """生成审核单唯一标识，格式: RV + 日期6位 + 随机4位hex，如 RV250501A3F1"""
    date_part = datetime.now().strftime("%y%m%d")
    rand_part = secrets.token_hex(2).upper()
    return f"{REVIEW_UID_PREFIX}{date_part}{rand_part}"


def _inject_review_uid_to_remark(order_data: dict[str, Any], review_uid: str) -> None:
    """将审核单UID标识注入到订单备注开头，格式: [RV250501A3F1] 原始备注"""
    if not review_uid:
        return
    tag = f"[{review_uid}]"
    existing_remark = str(order_data.get("remark") or "").strip()
    order_data["remark"] = f"{tag} {existing_remark}".strip()


def _check_review_uid_in_recent_orders(db: Session, erp_customer_id: str, review_uid: str) -> str | None:
    """检查该客户最近35张销售订单的备注中是否已包含该review_uid标识。
    返回匹配的订单号，未找到返回 None。"""
    if not review_uid or not erp_customer_id:
        return None
    tag = f"[{review_uid}]"
    rows = db.execute(
        text(
            "SELECT order_no, remark FROM erp_sales_orders "
            "WHERE customer_id = :cid ORDER BY order_date DESC, id DESC LIMIT 35"
        ),
        {"cid": erp_customer_id},
    ).mappings().all()
    for r in rows:
        remark = str(r.get("remark") or "")
        if tag in remark:
            return r["order_no"]
    return None


def _trigger_incremental_sync(order_no: str = "", product_nos: list[str] | None = None) -> None:
    """审核操作后在后台精准同步：只拉取刚下的那一单 + 涉及款号的库存，不阻塞当前请求。"""
    from app.services.erp_bridge import _erp_client
    if _erp_client is None:
        _logger.warning("[ReviewSync] ERPClient 未初始化，跳过增量同步")
        return

    from app.services.erp_sync import sync_single_order, sync_inventory_by_product_nos

    async def _do_sync():
        tasks = []
        if order_no:
            tasks.append(sync_single_order(_erp_client, order_no))
        if product_nos:
            tasks.append(sync_inventory_by_product_nos(_erp_client, product_nos))
        if not tasks:
            return
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for r in results:
            if isinstance(r, Exception):
                _logger.error("[ReviewSync] 精准同步任务异常: %s", r)

    try:
        loop = asyncio.get_running_loop()
        loop.create_task(_do_sync())
    except RuntimeError:
        _logger.warning("[ReviewSync] 无法获取事件循环，跳过增量同步")


class AttachmentContentRequiredError(Exception):
    pass


SUPPORTED_REVIEW_MESSAGE_TYPES = {"text", "image", "file"}


def _json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False)


def _json_loads(data: Any, default: Any):
    if not data:
        return default
    if isinstance(data, (dict, list)):
        return data
    try:
        return json.loads(data)
    except Exception:
        return default


def _find_first(data: Any, keys: list[str]) -> Any:
    if isinstance(data, dict):
        lowered = {str(key).lower(): value for key, value in data.items()}
        for key in keys:
            value = lowered.get(key.lower())
            if value not in (None, ""):
                return value
        for value in data.values():
            found = _find_first(value, keys)
            if found not in (None, ""):
                return found
    elif isinstance(data, list):
        for item in data:
            found = _find_first(item, keys)
            if found not in (None, ""):
                return found
    return None


def _safe_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (dict, list)):
        return _json_dumps(value)
    return str(value).strip()


def _normalize_instance_id(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _guess_attachment_mime(file_name: str, current_mime: str, message_type: str) -> str:
    if current_mime:
        return current_mime
    normalized_name = (file_name or "").lower()
    mime_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".bmp": "image/bmp",
        ".webp": "image/webp",
        ".gif": "image/gif",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".csv": "text/csv",
        ".txt": "text/plain",
        ".pdf": "application/pdf",
    }
    for ext, mime in mime_map.items():
        if normalized_name.endswith(ext):
            return mime
    if message_type == "image":
        return "image/png"
    return ""


def _read_local_attachment_base64(file_reference: str) -> str:
    if not file_reference or file_reference.startswith("data:") or "://" in file_reference:
        return ""
    try:
        path = Path(file_reference)
    except Exception:
        return ""
    if not path.is_file():
        return ""
    try:
        return base64.b64encode(path.read_bytes()).decode("ascii")
    except Exception:
        return ""


def _build_wechat_api_base_url(host: str, port: str) -> str:
    normalized_host = (host or "").strip()
    normalized_port = (port or "").strip()
    if not normalized_host:
        return ""
    if normalized_host.startswith("http://") or normalized_host.startswith("https://"):
        return normalized_host.rstrip("/") if not normalized_port else f"{normalized_host.rstrip('/')}:{normalized_port}"
    base = f"http://{normalized_host}"
    if normalized_port:
        base = f"{base}:{normalized_port}"
    return base.rstrip("/")


def _build_wechat_headers(api_key: str) -> dict[str, str]:
    headers = {}
    if api_key:
        headers["X-API-Key"] = api_key
    return headers


def _guess_attachment_extension(attachment_name: str, attachment_mime: str, message_type: str) -> str:
    normalized_name = (attachment_name or "").lower()
    known_exts = [
        ".png", ".jpg", ".jpeg", ".bmp", ".webp", ".gif", ".xlsx", ".xls", ".csv", ".txt", ".pdf", ".doc", ".docx"
    ]
    for ext in known_exts:
        if normalized_name.endswith(ext):
            return ext
    mime_map = {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/bmp": ".bmp",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/vnd.ms-excel": ".xls",
        "text/csv": ".csv",
        "text/plain": ".txt",
        "application/pdf": ".pdf",
    }
    if attachment_mime in mime_map:
        return mime_map[attachment_mime]
    return ".png" if message_type == "image" else ".bin"


def _infer_c2c_file_type(message_type: str, attachment_name: str, attachment_mime: str) -> int:
    if message_type == "image" or (attachment_mime or "").startswith("image/"):
        return 2
    normalized_name = (attachment_name or "").lower()
    if normalized_name.endswith((".mp4", ".mov", ".avi", ".mkv")):
        return 4
    if normalized_name.endswith((".mp3", ".wav", ".aac", ".amr")):
        return 3
    return 5


def _extract_download_request_data(payload: dict[str, Any], row: dict[str, Any]) -> dict[str, Any]:
    message = payload.get("message") if isinstance(payload.get("message"), dict) else {}
    message_data = message.get("data") if isinstance(message.get("data"), dict) else {}
    cdn_payload = message_data.get("cdn") if isinstance(message_data.get("cdn"), dict) else {}
    c2c_payload = message_data.get("c2c_cdn") if isinstance(message_data.get("c2c_cdn"), dict) else {}

    url = _find_first({"payload": payload, "message": message_data, "cdn": cdn_payload}, ["url"])
    auth_key = _find_first({"payload": payload, "message": message_data, "cdn": cdn_payload}, ["auth_key", "authKey"])
    aes_key = _find_first({"payload": payload, "message": message_data, "cdn": cdn_payload, "c2c": c2c_payload}, ["aes_key", "aesKey"])
    file_id = _find_first({"payload": payload, "message": message_data, "cdn": cdn_payload, "c2c": c2c_payload}, ["file_id", "fileId"])
    size = _find_first({"payload": payload, "message": message_data, "cdn": cdn_payload, "c2c": c2c_payload}, ["size", "file_size", "fileSize"])
    try:
        size = int(size or 0)
    except Exception:
        size = 0

    if url and auth_key and aes_key and size:
        return {
            "mode": "wx_download",
            "url": str(url),
            "auth_key": str(auth_key),
            "aes_key": str(aes_key),
            "size": size,
        }

    if file_id and aes_key:
        return {
            "mode": "c2c_download",
            "file_id": str(file_id),
            "aes_key": str(aes_key),
            "file_size": size,
            "file_type": _infer_c2c_file_type(str(row.get("message_type") or ""), str(row.get("attachment_name") or ""), str(row.get("attachment_mime") or "")),
        }

    return {}


def _resolve_wechat_runtime_config(db: Session, row: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    instance_id = row.get("instance_id")
    instance = None
    if instance_id:
        instance = db.query(WechatInstance).filter(WechatInstance.id == instance_id).first()
    if instance:
        return {
            "api_base_url": instance.api_base_url.rstrip("/"),
            "api_key": instance.api_key or "",
            "wxid": instance.wxid,
        }

    try:
        config_row = db.execute(text("SELECT host, port, api_key, selected_wxid FROM wechat_config WHERE id = 1")).mappings().first()
    except Exception:
        config_row = None

    payload_wxid = _safe_text(payload.get("wxid") or _find_first(payload, ["wxid", "robot_id", "robotId"]))
    if config_row:
        return {
            "api_base_url": _build_wechat_api_base_url(config_row.get("host") or "", config_row.get("port") or ""),
            "api_key": config_row.get("api_key") or "",
            "wxid": config_row.get("selected_wxid") or payload_wxid,
        }

    return {"api_base_url": "", "api_key": "", "wxid": payload_wxid}


async def _download_review_attachment_content(db: Session, row: dict[str, Any]) -> dict[str, Any]:
    payload = _json_loads(row.get("callback_payload"), {})
    runtime = _resolve_wechat_runtime_config(db, row, payload)
    if not runtime.get("api_base_url") or not runtime.get("wxid"):
        raise AttachmentContentRequiredError("缺少企业微信实例配置，无法下载附件")

    download_request = _extract_download_request_data(payload, row)
    if not download_request:
        raise AttachmentContentRequiredError("回调中缺少可用的 CDN 下载参数，无法下载附件")

    extension = _guess_attachment_extension(str(row.get("attachment_name") or ""), str(row.get("attachment_mime") or ""), str(row.get("message_type") or ""))
    download_dir = Path(__file__).resolve().parents[2] / "temp" / "wechat_attachments"
    download_dir.mkdir(parents=True, exist_ok=True)
    save_path = download_dir / f"review_{row['id']}{extension}"

    api_route = "cdn/wx_download" if download_request.get("mode") == "wx_download" else "cdn/c2c_download"
    request_body = dict(download_request)
    request_body["save_path"] = str(save_path)
    request_body.pop("mode", None)

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(
            f"{runtime['api_base_url']}/api/{runtime['wxid']}/{api_route}",
            json=request_body,
            headers=_build_wechat_headers(runtime.get("api_key") or ""),
        )
        response.raise_for_status()
        response_payload = response.json()

    if isinstance(response_payload, dict) and response_payload.get("code") not in (0, None):
        raise AttachmentContentRequiredError(response_payload.get("msg") or "企业微信附件下载失败")

    if not save_path.is_file():
        possible_path = _safe_text(_find_first(response_payload, ["save_path", "path", "file_path"]))
        if possible_path and Path(possible_path).is_file():
            save_path = Path(possible_path)
    if not save_path.is_file():
        raise AttachmentContentRequiredError("企业微信接口已返回成功，但未找到下载后的附件文件")

    file_bytes = save_path.read_bytes()
    attachment_base64 = base64.b64encode(file_bytes).decode("ascii")
    attachment_name = str(row.get("attachment_name") or save_path.name)
    attachment_mime = _guess_attachment_mime(attachment_name, str(row.get("attachment_mime") or ""), str(row.get("message_type") or ""))

    db.execute(
        text("UPDATE downstream_order_reviews SET attachment_name = :attachment_name, attachment_mime = :attachment_mime, attachment_base64 = :attachment_base64, parse_status = 'pending', ai_error = '', updated_at = NOW() WHERE id = :id"),
        {
            "id": row["id"],
            "attachment_name": attachment_name,
            "attachment_mime": attachment_mime,
            "attachment_base64": attachment_base64,
        },
    )
    db.commit()
    return {
        "attachment_name": attachment_name,
        "attachment_mime": attachment_mime,
        "save_path": str(save_path),
        "bytes": len(file_bytes),
    }


def _normalize_message_type(
    event_type: Any,
    content_type: Any,
    message_data: dict[str, Any],
    content_text: str,
    attachment_name: str,
    attachment_url: str,
    attachment_mime: str,
) -> str:
    event_key = _safe_text(event_type)
    if event_key.isdigit() and int(event_key) in EVENT_TYPE_TO_MESSAGE_TYPE:
        return EVENT_TYPE_TO_MESSAGE_TYPE[int(event_key)]

    content_key = _safe_text(content_type)
    if content_key.isdigit() and int(content_key) in CONTENT_TYPE_TO_MESSAGE_TYPE:
        return CONTENT_TYPE_TO_MESSAGE_TYPE[int(content_key)]

    if message_data.get("image") or attachment_mime.startswith("image/"):
        return "image"

    normalized_name = (attachment_name or attachment_url).lower()
    if normalized_name.endswith((".xlsx", ".xls", ".csv", ".txt", ".pdf", ".doc", ".docx", ".zip", ".rar", ".7z")):
        return "file"

    if any(message_data.get(key) for key in ["file_name", "file", "cdn", "c2c_cdn"]):
        return "file"

    if content_text:
        return "text"

    return _safe_text(message_data.get("message_type") or event_type or "unknown").lower() or "unknown"


def _extract_callback_message(payload: dict[str, Any], instance_id: Optional[str]) -> dict[str, Any]:
    payload = payload if isinstance(payload, dict) else {}

    # 兼容三种回调格式:
    # 1. NGCBotV3-QW 转发格式: { "message": { "type": ..., "data": {...} }, "wxid": "..." }
    # 2. 原始 API 格式:        { "type": 11041, "data": { "content": "...", ... } }
    # 3. NGCDemo legacy 格式:   { "message": { "type": ..., "data": { "msg": "...", "from_wxid": "..." } } }
    message = payload.get("message") if isinstance(payload.get("message"), dict) else {}
    message_data = message.get("data") if isinstance(message.get("data"), dict) else {}

    # 原始 API 格式: type 和 data 在顶层，没有 message 包裹
    if not message_data and isinstance(payload.get("data"), dict) and payload.get("type"):
        message_data = payload["data"]

    source_format = "generic"
    if message_data and any(key in message_data for key in ["conversation_id", "sender", "sender_name", "content_type"]):
        source_format = "ngcbot_v3"
    elif message_data and any(key in message_data for key in ["room_wxid", "from_wxid", "msg", "wx_type"]):
        source_format = "ngcdemo_legacy"

    normalized_instance_id = _normalize_instance_id(instance_id or payload.get("instanceId") or payload.get("instance_id"))
    event_type = message.get("type") or payload.get("type") or payload.get("message_type") or payload.get("msg_type")
    content_type = message_data.get("content_type") or payload.get("content_type") or message_data.get("wx_type")

    sender_id = _safe_text(
        message_data.get("sender")
        or message_data.get("from_wxid")
        or payload.get("sender_id")
        or payload.get("from_wxid")
        or payload.get("sender")
    )
    sender_name = _safe_text(
        message_data.get("sender_name")
        or message_data.get("from_name")
        or payload.get("sender_name")
        or payload.get("nickname")
    )

    room_id = _safe_text(
        message_data.get("room_wxid")
        or message_data.get("room_conversation_id")
        or payload.get("room_id")
        or payload.get("roomid")
    )
    conversation_id = _safe_text(message_data.get("conversation_id") or payload.get("conversation_id"))
    if not room_id and conversation_id and sender_id and conversation_id != sender_id:
        room_id = conversation_id

    room_name = _safe_text(
        message_data.get("room_name")
        or message_data.get("conversation_name")
        or payload.get("room_name")
        or payload.get("conversation_name")
    )

    content_text = _safe_text(
        message_data.get("content")
        or message_data.get("text_content")
        or message_data.get("msg")
        or payload.get("content")
        or payload.get("msg")
        or payload.get("text")
    )

    cdn_payload = message_data.get("cdn") if isinstance(message_data.get("cdn"), dict) else {}
    c2c_cdn_payload = message_data.get("c2c_cdn") if isinstance(message_data.get("c2c_cdn"), dict) else {}
    media_payload = cdn_payload or c2c_cdn_payload

    local_attachment_ref = _safe_text(
        message_data.get("image")
        or message_data.get("file")
        or message_data.get("video")
        or payload.get("file_path")
        or payload.get("image")
    )
    attachment_name = _safe_text(
        message_data.get("file_name")
        or media_payload.get("file_name")
        or payload.get("file_name")
        or payload.get("filename")
        or payload.get("name")
        or payload.get("title")
    )
    if local_attachment_ref and not attachment_name:
        try:
            attachment_name = Path(local_attachment_ref).name
        except Exception:
            attachment_name = ""

    attachment_url = _safe_text(
        payload.get("file_url")
        or payload.get("download_url")
        or payload.get("downloadUrl")
        or media_payload.get("url")
        or local_attachment_ref
    )
    attachment_mime = _safe_text(
        payload.get("mime_type")
        or payload.get("mimetype")
        or payload.get("content_type")
        or message_data.get("mime_type")
    )
    attachment_base64 = _safe_text(
        payload.get("file_base64")
        or payload.get("base64")
        or payload.get("data_base64")
        or message_data.get("file_base64")
    )

    if not attachment_base64 and local_attachment_ref:
        attachment_base64 = _read_local_attachment_base64(local_attachment_ref)

    if attachment_base64.startswith("data:"):
        header, encoded = attachment_base64.split(",", 1)
        attachment_base64 = encoded
        if not attachment_mime and ";base64" in header:
            attachment_mime = header.split(":", 1)[1].split(";", 1)[0]

    message_type = _normalize_message_type(
        event_type,
        content_type,
        message_data,
        content_text,
        attachment_name,
        attachment_url,
        attachment_mime,
    )
    attachment_mime = _guess_attachment_mime(attachment_name or attachment_url, attachment_mime, message_type)

    if not content_text and message_type in {"image", "file"}:
        prefix = "图片" if message_type == "image" else "文件"
        content_text = f"[{prefix}] {attachment_name or attachment_url or '附件消息'}"

    is_group_message = bool(room_id) and (room_id != sender_id or not sender_id)
    has_usable_content = bool(content_text or attachment_base64 or attachment_url or attachment_name)
    requires_attachment_download = message_type in {"image", "file"} and not attachment_base64 and bool(attachment_url or attachment_name)

    skip_reason = ""
    if not payload:
        skip_reason = "empty_payload"
    elif not is_group_message:
        skip_reason = "not_group_message"
    elif message_type not in SUPPORTED_REVIEW_MESSAGE_TYPES:
        skip_reason = f"unsupported_message_type:{message_type}"
    elif not has_usable_content:
        skip_reason = "empty_message_content"

    return {
        "source_type": f"wechat_{source_format}",
        "instance_id": normalized_instance_id,
        "room_id": room_id,
        "room_name": room_name,
        "sender_id": sender_id,
        "sender_name": sender_name,
        "message_type": message_type,
        "content_text": content_text,
        "attachment_name": attachment_name,
        "attachment_url": attachment_url,
        "attachment_mime": attachment_mime,
        "attachment_base64": attachment_base64,
        "event_type": _safe_text(event_type),
        "content_type": _safe_text(content_type),
        "source_format": source_format,
        "is_group_message": is_group_message,
        "requires_attachment_download": requires_attachment_download,
        "is_order_candidate": not skip_reason,
        "skip_reason": skip_reason,
    }


def _extract_excel_summary(attachment_base64: str) -> str:
    if not attachment_base64:
        return ""
    if load_workbook is None:
        return "Excel 文件已收到，但当前环境未安装 openpyxl。"
    try:
        binary = base64.b64decode(attachment_base64)
        workbook = load_workbook(io.BytesIO(binary), data_only=True)
    except Exception as exc:
        return f"Excel 解析失败: {exc}"

    lines: list[str] = []
    for sheet in workbook.worksheets[:3]:
        lines.append(f"[Sheet] {sheet.title}")
        for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
            row_values = ["" if cell is None else str(cell).strip() for cell in row[:20]]
            if any(row_values):
                lines.append("\t".join(row_values))
    return "\n".join(lines[:240])


def _resolve_product_no_for_context(db: Session, name: str) -> str:
    """通过映射表或产品表解析 AI 提取的款号 → 真实货号。

    优先级：映射表 > 产品表精确匹配 > 原值
    这样即使 AI 提取的款号在产品表中存在，只要映射表里配了别名也会被正确映射。
    """
    name = name.strip()
    if not name:
        return name
    # 1. 优先查映射表——用户显式配置的别名拥有最高优先级
    try:
        alias = db.execute(
            text("SELECT product_no FROM product_name_mappings WHERE alias_name = :name LIMIT 1"),
            {"name": name},
        ).mappings().first()
        if alias:
            return alias["product_no"]
    except Exception:
        pass
    # 2. 精确匹配 erp_products
    direct = db.execute(
        text("SELECT product_no FROM erp_products WHERE product_no = :name LIMIT 1"),
        {"name": name},
    ).mappings().first()
    if direct:
        return direct["product_no"]
    return name


def query_product_context_for_nos(db: Session, product_nos: list[str]) -> str:
    """根据款号列表查询产品表中每个款号的可选颜色和尺码，返回格式化文本。

    流程：
    1. 先通过产品表精确匹配或映射表解析别名 → 真实货号
    2. 用真实货号查询 erp_products 获取可选颜色（color 字段）和尺码规格（spec 字段）
    """
    if not product_nos:
        return "未提取到款号，无法查询产品信息。"

    from app.services.erp_sync import ensure_tables
    ensure_tables(db)

    lines: list[str] = []
    for pno in product_nos:
        # 步骤1：解析真实货号（精确匹配 + 别名映射）
        resolved_pno = _resolve_product_no_for_context(db, pno)
        pno_label = f"{pno}（→{resolved_pno}）" if resolved_pno != pno else pno

        # 步骤2：从产品表查询颜色和尺码规格
        product_row = db.execute(
            text("SELECT color, spec FROM erp_products WHERE product_no = :pno LIMIT 1"),
            {"pno": resolved_pno},
        ).mappings().first()

        if not product_row:
            lines.append(f"款号 {pno_label}：产品表中未找到该货号")
            continue

        color_text = (product_row["color"] or "").strip()
        spec_text = (product_row["spec"] or "").strip()

        if color_text or spec_text:
            lines.append(f"款号 {pno_label} 可选信息：")
            if color_text:
                lines.append(f"  可选颜色：{color_text}")
            if spec_text:
                lines.append(f"  可选尺码：{spec_text}")
        else:
            lines.append(f"款号 {pno_label}：产品表中无颜色和尺码信息")

    return "\n".join(lines)


def query_current_year_catalog(db: Session) -> list[dict[str, Any]]:
    """查询所有本年产品库的产品号及其名称映射、可选颜色和尺码，供 AI 款号匹配和上下文解析使用。

    返回:
        [
            {
                "product_no": "1234",
                "product_name": "...",
                "aliases": ["别名1", "别名2"],
                "color": "黑色,白色",
                "spec": "M,L,XL",
            },
            ...
        ]
    """
    from app.services.erp_sync import ensure_tables
    ensure_tables(db)

    rows = db.execute(
        text("SELECT product_no, product_name, color, spec FROM erp_products WHERE is_current_year = 1 ORDER BY product_no"),
    ).mappings().all()

    catalog: list[dict[str, Any]] = []
    pno_set: set[str] = set()
    for r in rows:
        pno = (r["product_no"] or "").strip()
        if not pno or pno in pno_set:
            continue
        pno_set.add(pno)
        catalog.append({
            "product_no": pno,
            "product_name": (r["product_name"] or "").strip(),
            "aliases": [],
            "color": (r["color"] or "").strip(),
            "spec": (r["spec"] or "").strip(),
        })

    # 加载映射关系
    try:
        mapping_rows = db.execute(
            text("SELECT product_no, alias_name FROM product_name_mappings ORDER BY product_no"),
        ).mappings().all()
        alias_map: dict[str, list[str]] = {}
        for mr in mapping_rows:
            target = (mr["product_no"] or "").strip()
            alias = (mr["alias_name"] or "").strip()
            if target and alias:
                alias_map.setdefault(target, []).append(alias)
        for item in catalog:
            item["aliases"] = alias_map.get(item["product_no"], [])
    except Exception:
        pass

    return catalog


def build_context_from_catalog(
    catalog: list[dict[str, Any]],
    product_nos: list[str],
) -> dict[str, Any]:
    """从已加载的 catalog 中提取指定款号的可选颜色/尺码/映射，无需再查 DB。

    返回格式与 query_product_context_structured 一致。
    """
    # 构建 catalog 查找表
    catalog_map: dict[str, dict[str, Any]] = {item["product_no"]: item for item in catalog}

    # 收集所有映射关系
    mappings: dict[str, str] = {}
    for item in catalog:
        for alias in item.get("aliases") or []:
            mappings[alias] = item["product_no"]

    products: dict[str, dict[str, list[str]]] = {}
    all_sizes: list[str] = []
    all_colors: list[str] = []
    seen_sizes: set[str] = set()
    seen_colors: set[str] = set()

    for pno in product_nos:
        # 先解析别名
        resolved_pno = mappings.get(pno, pno)
        item = catalog_map.get(resolved_pno)
        if not item:
            continue

        pno_colors: list[str] = []
        for c in (item.get("color") or "").split(","):
            c = c.strip()
            if c:
                pno_colors.append(c)
                if c not in seen_colors:
                    seen_colors.add(c)
                    all_colors.append(c)

        pno_sizes: list[str] = []
        for s in (item.get("spec") or "").split(","):
            s = s.strip()
            if s:
                pno_sizes.append(s)
                if s not in seen_sizes:
                    seen_sizes.add(s)
                    all_sizes.append(s)

        products[pno] = {"sizes": pno_sizes, "colors": pno_colors}

    return {"products": products, "sizes": all_sizes, "colors": all_colors, "mappings": mappings}


def query_product_context_structured(db: Session, product_nos: list[str]) -> dict[str, Any]:
    """根据款号列表查询产品表，返回按款号分组的可选尺码、颜色和款号映射。

    返回:
        {
            "products": {
                "1234": {"sizes": ["M", "L", "XL"], "colors": ["黑色", "白色"]},
                "5678": {"sizes": ["S", "M"], "colors": ["红色"]},
            },
            "sizes": ["M", "L", "XL", "S", ...],       # 所有款号的尺码并集（向后兼容）
            "colors": ["黑色", "白色", "红色", ...],    # 所有款号的颜色并集（向后兼容）
            "mappings": {"原始款号": "目标款号", ...},
        }
    """
    from app.services.erp_sync import ensure_tables
    ensure_tables(db)

    products: dict[str, dict[str, list[str]]] = {}
    all_sizes: list[str] = []
    all_colors: list[str] = []
    seen_sizes: set[str] = set()
    seen_colors: set[str] = set()

    for pno in (product_nos or []):
        resolved_pno = _resolve_product_no_for_context(db, pno)
        product_row = db.execute(
            text("SELECT color, spec FROM erp_products WHERE product_no = :pno LIMIT 1"),
            {"pno": resolved_pno},
        ).mappings().first()
        if not product_row:
            continue
        pno_colors: list[str] = []
        pno_sizes: list[str] = []
        for c in (product_row["color"] or "").split(","):
            c = c.strip()
            if c:
                pno_colors.append(c)
                if c not in seen_colors:
                    seen_colors.add(c)
                    all_colors.append(c)
        for s in (product_row["spec"] or "").split(","):
            s = s.strip()
            if s:
                pno_sizes.append(s)
                if s not in seen_sizes:
                    seen_sizes.add(s)
                    all_sizes.append(s)
        products[pno] = {"sizes": pno_sizes, "colors": pno_colors}

    # 款号映射：alias_name（图片中原始款号） → product_no（目标款号）
    mappings: dict[str, str] = {}
    try:
        rows = db.execute(
            text("SELECT product_no, alias_name FROM product_name_mappings ORDER BY product_no"),
        ).mappings().all()
        for r in rows:
            alias = (r["alias_name"] or "").strip()
            target = (r["product_no"] or "").strip()
            if alias and target:
                mappings[alias] = target
    except Exception:
        pass

    return {"products": products, "sizes": all_sizes, "colors": all_colors, "mappings": mappings}


def _normalize_discount(value: Any) -> int:
    try:
        number = float(value)
    except Exception:
        return 100
    if number <= 1:
        return int(number * 100)
    return int(number)


def _normalize_order(parsed: dict[str, Any], customer_name: str = "") -> dict[str, Any]:
    items = []
    for item in parsed.get("items") or []:
        sizes = []
        for size in item.get("sizes") or []:
            name = str(size.get("size") or "").strip()
            qty = int(size.get("qty") or 0)
            if name and qty:
                sizes.append({"size": name, "qty": qty})
        if not sizes:
            continue
        items.append({
            "product_no": str(item.get("product_no") or "").strip(),
            "color": str(item.get("color") or "").strip(),
            "sizes": sizes,
            "remark": str(item.get("remark") or "").strip(),
        })

    return {
        "customer_name": str(customer_name or "").strip(),
        "order_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "remark": str(parsed.get("remark") or "").strip(),
        "items": items,
        "uncertainties": parsed.get("uncertainties") or [],
    }


def ensure_review_state(db: Session):
    ensure_downstream_support_tables(db)


def resolve_customer_by_room(db: Session, room_id: str, instance_id: Optional[str] = None) -> Optional[dict[str, Any]]:
    if not room_id:
        return None
    # 兼容 room_id 有/无 'R:' 前缀（回调消息带前缀，前端保存不带前缀）
    rid_clean = room_id[2:] if room_id.startswith("R:") else room_id
    params: dict[str, Any] = {"rid1": rid_clean, "rid2": f"R:{rid_clean}"}
    sql = (
        "SELECT c.id, c.customer_name, c.contact_person, c.phone, c.address, c.erp_customer_id, r.instance_id, r.room_id, r.room_name "
        "FROM downstream_customer_wechat_rooms r "
        "INNER JOIN downstream_customers c ON c.id = r.customer_id "
        "WHERE r.room_id IN (:rid1, :rid2) AND c.deleted_at IS NULL AND c.status = 1"
    )
    if instance_id:
        try:
            params["instance_id"] = int(instance_id)
            sql += " AND (r.instance_id = :instance_id OR r.instance_id IS NULL)"
        except (ValueError, TypeError):
            pass
    sql += " ORDER BY c.id ASC LIMIT 1"
    result = db.execute(text(sql), params).mappings().first()
    if result is None:
        _log = logging.getLogger(__name__)
        _log.info("resolve_customer_by_room: 未匹配 room_id=%r(%r) instance_id=%r", room_id, rid_clean, instance_id)
    return result


def _lookup_room_names(db: Session, room_ids: list[str]) -> dict[str, str]:
    """从 API 缓存和 wechat_room_listeners 批量查找 room_id → room_name 映射"""
    if not room_ids:
        return {}
    unique_ids = list(set(rid for rid in room_ids if rid))
    if not unique_ids:
        return {}

    # 优先从 API 缓存获取
    result: dict[str, str] = {}
    try:
        from app.services.wechat_room_cache import get_room_names
        result.update(get_room_names(unique_ids))
    except Exception:
        pass

    # 缓存未命中的，从 DB 查找
    missing = [rid for rid in unique_ids if rid not in result]
    if missing:
        placeholders = ", ".join(f":rid{i}" for i in range(len(missing)))
        params = {f"rid{i}": rid for i, rid in enumerate(missing)}
        try:
            rows = db.execute(
                text(f"SELECT room_id, room_name FROM wechat_room_listeners WHERE room_id IN ({placeholders})"),
                params,
            ).mappings().all()
            for r in rows:
                if r.get("room_name"):
                    result[r["room_id"]] = r["room_name"]
        except Exception:
            pass

    return result


def serialize_review(row: dict[str, Any]) -> dict[str, Any]:
    item = dict(row)
    item["callback_payload"] = _json_loads(item.get("callback_payload"), {})
    item["parsed_order"] = _json_loads(item.get("parsed_order_json"), None)
    item["manual_order"] = _json_loads(item.get("manual_order_json"), None)
    item["replace_source_rows"] = _json_loads(item.get("replace_source_ids"), [])
    item.pop("parsed_order_json", None)
    item.pop("manual_order_json", None)
    item.pop("replace_source_ids", None)
    return item


def _mask_secret(value: str) -> str:
    text_value = str(value or "")
    if len(text_value) <= 8:
        return "*" * len(text_value)
    return f"{text_value[:4]}***{text_value[-4:]}"


def get_review_attachment_debug(db: Session, review_id: int) -> dict[str, Any]:
    ensure_review_state(db)
    row = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")

    payload = _json_loads(row.get("callback_payload"), {})
    runtime = _resolve_wechat_runtime_config(db, row, payload)
    download_request = _extract_download_request_data(payload, row)

    masked_download_request = dict(download_request)
    for sensitive_key in ["auth_key", "aes_key", "file_id"]:
        if sensitive_key in masked_download_request and masked_download_request[sensitive_key]:
            masked_download_request[sensitive_key] = _mask_secret(masked_download_request[sensitive_key])

    return {
        "review_id": review_id,
        "source_type": row.get("source_type") or "",
        "message_type": row.get("message_type") or "",
        "parse_status": row.get("parse_status") or "",
        "review_status": row.get("review_status") or "",
        "attachment_name": row.get("attachment_name") or "",
        "attachment_url": row.get("attachment_url") or "",
        "attachment_mime": row.get("attachment_mime") or "",
        "has_attachment_base64": bool(row.get("attachment_base64")),
        "attachment_base64_length": len(str(row.get("attachment_base64") or "")),
        "ai_error": row.get("ai_error") or "",
        "runtime": {
            "api_base_url": runtime.get("api_base_url") or "",
            "wxid": runtime.get("wxid") or "",
            "api_key_configured": bool(runtime.get("api_key")),
        },
        "download": {
            "ready": bool(runtime.get("api_base_url") and runtime.get("wxid") and download_request),
            "mode": download_request.get("mode") or "",
            "request": masked_download_request,
        },
        "callback_payload": payload,
    }


async def retry_review_attachment_download(db: Session, review_id: int, *, reparse: bool = True, force_download: bool = False) -> dict[str, Any]:
    ensure_review_state(db)
    row = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")

    message_type = str(row.get("message_type") or "").lower()
    if message_type not in {"image", "file"}:
        raise ValueError("当前记录不是需要下载附件的图片或文件消息")

    download_result = None
    if force_download or not row.get("attachment_base64"):
        download_result = await _download_review_attachment_content(db, row)
    else:
        download_result = {
            "skipped": True,
            "reason": "attachment_already_exists",
            "attachment_name": row.get("attachment_name") or "",
            "attachment_mime": row.get("attachment_mime") or "",
        }

    parsed_order = None
    parse_error = ""
    if reparse:
        try:
            parsed_order = await parse_review_content(db, review_id)
        except Exception as exc:
            parse_error = str(exc)

    latest = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    response = {
        "download": download_result,
        "review": serialize_review(latest),
    }
    if parsed_order:
        response["parsed_order"] = parsed_order
    if parse_error:
        response["parse_error"] = parse_error
    return response


async def parse_review_content(db: Session, review_id: int) -> dict[str, Any]:
    ensure_review_state(db)
    row = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")

    if str(row.get("message_type") or "").lower() in {"image", "file"} and not row.get("attachment_base64"):
        await _download_review_attachment_content(db, row)
        row = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()

    attachment_name = str(row.get("attachment_name") or "").lower()
    attachment_mime = str(row.get("attachment_mime") or "").lower()
    message_type = str(row.get("message_type") or "text").lower()
    customer_hint = str(row.get("customer_name") or row.get("room_name") or "")

    try:
        if message_type in {"image", "file"} and not row.get("attachment_base64"):
            raise AttachmentContentRequiredError("当前消息仅收到附件元数据，请先下载企业微信附件内容后再解析")

        # 构建统一消息列表供 AI 使用
        is_excel = ("excel" in attachment_mime or attachment_name.endswith(".xlsx") or attachment_name.endswith(".xls")) and row.get("attachment_base64")
        is_image = (message_type in {"image", "img", "picture"} or attachment_mime.startswith("image/")) and row.get("attachment_base64")

        context_messages: list[dict[str, Any]] = []
        excel_summary = ""
        if is_excel:
            excel_summary = _extract_excel_summary(row.get("attachment_base64") or "")
            context_messages.append({
                "type": "file",
                "file_name": row.get("attachment_name") or "",
                "excel_summary": excel_summary,
                "content": excel_summary,
            })
        elif is_image:
            context_messages.append({
                "type": "image",
                "base64": row.get("attachment_base64") or "",
                "mime": row.get("attachment_mime") or "image/png",
            })
            if row.get("content_text"):
                context_messages.insert(0, {"type": "text", "content": row.get("content_text")})
        else:
            text_content = row.get("content_text") or row.get("attachment_name") or row.get("room_name") or ""
            context_messages.append({"type": "text", "content": text_content})

        # === 步骤 1：智能体 A — 从本年产品目录匹配款号 + 判断旋转角度 ===
        logger.info("[AI Parse] review=%d 步骤1: 加载本年产品目录并匹配款号...", review_id)
        catalog = query_current_year_catalog(db)
        logger.info("[AI Parse] review=%d 本年产品目录: %d 个产品", review_id, len(catalog))
        extract_result = await ai_order_parser.extract_product_nos(context_messages, db=db, catalog=catalog)
        product_nos = extract_result.get("product_nos") or []
        no_match = extract_result.get("no_match", not product_nos)
        rotation_angle = extract_result.get("rotation_angle") or 0
        logger.info("[AI Parse] review=%d 步骤1: product_nos=%s no_match=%s rotation=%d°", review_id, product_nos, no_match, rotation_angle)

        # === 步骤 1 校验：款号必须在本年产品目录中，否则中止 ===
        if catalog and not product_nos:
            err_msg = "客户提到的内容在本年产品目录中没有匹配项" if no_match else "未识别到任何有效款号信息"
            logger.warning("[AI Parse] review=%d 步骤1未匹配到有效款号，中止解析 no_match=%s", review_id, no_match)
            db.execute(
                text("UPDATE downstream_order_reviews SET parse_status = 'failed', ai_error = :err, updated_at = NOW() WHERE id = :id"),
                {"id": review_id, "err": err_msg},
            )
            db.commit()
            raise AIOrderParserError(err_msg)

        # === 步骤 1.5：根据 AI 判断的角度旋转图片 ===
        if rotation_angle and rotation_angle != 0:
            from app.services.ai_order_parser import rotate_images_in_messages
            context_messages = rotate_images_in_messages(context_messages, rotation_angle)
            logger.info("[AI Parse] review=%d 图片已旋转 %d°", review_id, rotation_angle)

        # === 步骤 2：从商品列表（catalog）获取可选颜色/尺码/映射 ===
        context_data = build_context_from_catalog(catalog, product_nos) if product_nos else {}
        logger.info("[AI Parse] review=%d 步骤2: sizes=%s colors=%d mappings=%d",
                     review_id, context_data.get("sizes"), len(context_data.get("colors", [])),
                     len(context_data.get("mappings", {})))

        # === 步骤 3：智能体 B — 带上下文解析完整订单 ===
        logger.info("[AI Parse] review=%d 步骤3: 带上下文解析订单...", review_id)
        if context_data and context_data.get("products"):
            parsed = await ai_order_parser.parse_with_product_context(
                context_messages, context_data, customer_hint=customer_hint, db=db,
            )
        else:
            # 没有提取到款号时，回退到批量解析（也带目录约束）
            parsed = await ai_order_parser.parse_batch(context_messages, customer_hint=customer_hint, db=db, catalog=catalog)

        normalized = _normalize_order(parsed, customer_hint)

        # 硬校验：过滤掉不在本年产品目录中的款号
        if catalog:
            valid_pnos = {item["product_no"] for item in catalog}
            original_items = normalized.get("items") or []
            filtered_items = [it for it in original_items
                              if it.get("product_no") and it["product_no"] in valid_pnos]
            removed_pnos = {it.get("product_no") for it in original_items
                            if it.get("product_no") and it["product_no"] not in valid_pnos}
            if removed_pnos:
                logger.warning("[AI Parse] review=%d 以下款号不在本年产品目录中(已过滤): %s",
                               review_id, removed_pnos)
            normalized["items"] = filtered_items

            if not filtered_items:
                hint = "、".join(sorted(removed_pnos)[:5]) if removed_pnos else ""
                err_msg = f"解析到的款号 [{hint}] 均不在本年产品目录中，非我司产品" if hint else "未解析到有效的产品款号"
                logger.warning("[AI Parse] review=%d 步骤3后所有款号均不在目录中，中止", review_id)
                db.execute(
                    text("UPDATE downstream_order_reviews SET parse_status = 'failed', ai_error = :err, updated_at = NOW() WHERE id = :id"),
                    {"id": review_id, "err": err_msg},
                )
                db.commit()
                raise AIOrderParserError(err_msg)

        db.execute(
            text("UPDATE downstream_order_reviews SET parse_status = 'success', ai_error = '', parsed_order_json = :parsed_order_json, updated_at = NOW() WHERE id = :id"),
            {"id": review_id, "parsed_order_json": _json_dumps(normalized)},
        )
        db.commit()
        logger.info("[AI Parse] review=%d 解析完成, items=%d", review_id, len(normalized.get("items") or []))
        return normalized
    except AttachmentContentRequiredError as exc:
        db.execute(
            text("UPDATE downstream_order_reviews SET parse_status = 'pending_attachment', ai_error = :ai_error, updated_at = NOW() WHERE id = :id"),
            {"id": review_id, "ai_error": str(exc)},
        )
        db.commit()
        raise
    except AIOrderParserError as exc:
        db.execute(
            text("UPDATE downstream_order_reviews SET parse_status = 'failed', ai_error = :ai_error, updated_at = NOW() WHERE id = :id"),
            {"id": review_id, "ai_error": str(exc)},
        )
        db.commit()
        raise


async def create_review_from_callback(db: Session, payload: dict[str, Any], instance_id: Optional[str] = None) -> dict[str, Any]:
    ensure_review_state(db)
    message = _extract_callback_message(payload, instance_id)
    if not message.get("is_order_candidate"):
        return {
            "skipped": True,
            "reason": message.get("skip_reason") or "not_order_candidate",
            "message_type": message.get("message_type") or "",
            "room_id": message.get("room_id") or "",
            "source_format": message.get("source_format") or "",
        }

    customer = resolve_customer_by_room(db, message["room_id"], message["instance_id"])
    initial_parse_status = "pending_attachment" if message.get("requires_attachment_download") else "pending"

    # 如果 callback 中没有 room_name，从 wechat_room_listeners 补全
    room_name = message.get("room_name") or ""
    if not room_name and message.get("room_id"):
        room_map = _lookup_room_names(db, [message["room_id"]])
        room_name = room_map.get(message["room_id"], "")

    review_uid = _generate_review_uid()
    result = db.execute(
        text(
            "INSERT INTO downstream_order_reviews ("
            "review_uid, source_type, instance_id, room_id, room_name, sender_id, sender_name, message_type, content_text, attachment_name, attachment_url, attachment_mime, attachment_base64, callback_payload, parse_status, review_status, customer_id, customer_name"
            ") VALUES ("
            ":review_uid, :source_type, :instance_id, :room_id, :room_name, :sender_id, :sender_name, :message_type, :content_text, :attachment_name, :attachment_url, :attachment_mime, :attachment_base64, :callback_payload, :parse_status, 'pending', :customer_id, :customer_name"
            ")"
        ),
        {
            "review_uid": review_uid,
            "source_type": message.get("source_type") or "wechat_generic",
            "instance_id": message.get("instance_id"),
            "room_id": message.get("room_id") or "",
            "room_name": room_name,
            "sender_id": message.get("sender_id") or "",
            "sender_name": message.get("sender_name") or "",
            "message_type": message.get("message_type") or "text",
            "content_text": message.get("content_text") or "",
            "attachment_name": message.get("attachment_name") or "",
            "attachment_url": message.get("attachment_url") or "",
            "attachment_mime": message.get("attachment_mime") or "",
            "attachment_base64": message.get("attachment_base64") or "",
            "callback_payload": _json_dumps(payload),
            "parse_status": initial_parse_status,
            "customer_id": customer["id"] if customer else None,
            "customer_name": customer["customer_name"] if customer else "",
        },
    )
    review_id = result.lastrowid
    db.commit()

    parsed_order = None
    parse_error = ""
    try:
        parsed_order = await parse_review_content(db, review_id)
    except Exception as exc:
        parse_error = str(exc)

    latest = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    response = serialize_review(latest)
    if parse_error:
        response["parse_error"] = parse_error
    if parsed_order:
        response["parsed_order"] = parsed_order
    from app.services.review_events import notify_review_change
    notify_review_change("new_review", {"review_id": review_id})
    return response


def list_reviews(db: Session, page: int = 1, page_size: int = 20, review_status: str = "", customer_id: Optional[int] = None, sort: Optional[str] = None) -> dict[str, Any]:
    ensure_review_state(db)
    params = {"limit": page_size, "offset": (page - 1) * page_size}
    where_parts = ["1 = 1"]
    if review_status:
        where_parts.append("r.review_status = :review_status")
        params["review_status"] = review_status
    # 待审核列表只显示AI解析成功的记录
    if review_status == "pending" or not review_status:
        where_parts.append("r.parse_status = 'success'")
    if customer_id:
        where_parts.append("r.customer_id = :customer_id")
        params["customer_id"] = customer_id
    where_sql = " AND ".join(where_parts)
    # 排序：asc=从早到晚，desc=从晚到早；默认待审核按 ASC，其余按 DESC
    order_dir = "ASC" if sort == "asc" else ("DESC" if sort == "desc" else ("ASC" if review_status == "pending" else "DESC"))
    rows = db.execute(
        text(
            f"SELECT r.*, "
            f"COALESCE(NULLIF(r.room_name, ''), m.room_name, '') AS _room_name, "
            f"COALESCE(NULLIF(r.sender_name, ''), m.sender_name, '') AS _sender_name, "
            f"COALESCE(NULLIF(r.message_type, ''), m.message_type, 'text') AS _message_type "
            f"FROM downstream_order_reviews r "
            f"LEFT JOIN message_logs m ON r.msg_log_id IS NOT NULL AND r.msg_log_id = m.id "
            f"WHERE {where_sql} ORDER BY r.created_at {order_dir} LIMIT :limit OFFSET :offset"
        ),
        params,
    ).mappings().all()
    count_params = {key: value for key, value in params.items() if key not in {"limit", "offset"}}
    total = db.execute(text(f"SELECT COUNT(*) AS total FROM downstream_order_reviews r WHERE {where_sql}"), count_params).mappings().first()["total"]

    # 批量查找缺失的 room_name
    missing_room_ids = []
    serialized = []
    for row in rows:
        item = serialize_review(row)
        # 用 JOIN 得到的补全值覆盖空字段
        if not item.get("room_name"):
            item["room_name"] = row.get("_room_name") or ""
        if not item.get("sender_name"):
            item["sender_name"] = row.get("_sender_name") or ""
        if not item.get("message_type") or item["message_type"] in ("text", "unknown"):
            enriched_type = row.get("_message_type") or ""
            if enriched_type and enriched_type != "unknown":
                item["message_type"] = enriched_type
        if not item.get("room_name") and item.get("room_id"):
            missing_room_ids.append(item["room_id"])
        serialized.append(item)

    if missing_room_ids:
        room_map = _lookup_room_names(db, missing_room_ids)
        for item in serialized:
            if not item.get("room_name") and item.get("room_id") and item["room_id"] in room_map:
                item["room_name"] = room_map[item["room_id"]]

    return {
        "list": serialized,
        "total": total,
        "page": page,
        "pageSize": page_size,
    }


def get_review_detail(db: Session, review_id: int) -> dict[str, Any]:
    ensure_review_state(db)
    row = db.execute(
        text(
            "SELECT r.*, "
            "COALESCE(NULLIF(r.room_name, ''), m.room_name, '') AS _room_name, "
            "COALESCE(NULLIF(r.sender_name, ''), m.sender_name, '') AS _sender_name, "
            "COALESCE(NULLIF(r.message_type, ''), m.message_type, 'text') AS _message_type "
            "FROM downstream_order_reviews r "
            "LEFT JOIN message_logs m ON r.msg_log_id IS NOT NULL AND r.msg_log_id = m.id "
            "WHERE r.id = :id"
        ),
        {"id": review_id},
    ).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")
    item = serialize_review(row)
    if not item.get("room_name"):
        item["room_name"] = row.get("_room_name") or ""
    if not item.get("sender_name"):
        item["sender_name"] = row.get("_sender_name") or ""
    if not item.get("message_type") or item["message_type"] in ("text", "unknown"):
        enriched_type = row.get("_message_type") or ""
        if enriched_type and enriched_type != "unknown":
            item["message_type"] = enriched_type
    # 从 wechat_room_listeners 补全 room_name
    if not item.get("room_name") and item.get("room_id"):
        room_map = _lookup_room_names(db, [item["room_id"]])
        if item["room_id"] in room_map:
            item["room_name"] = room_map[item["room_id"]]
    return item


def get_review_context_messages(db: Session, review_id: int) -> dict[str, Any]:
    """获取审核记录关联的上下文消息（触发消息前20条+后20条），用于聊天记录展示。
    图片消息会关联查询审核记录中的 attachment_base64 以便前端直接展示。"""
    from app.services.message_logs import ensure_message_logs_table
    ensure_message_logs_table(db)
    ensure_review_state(db)

    row = db.execute(
        text("SELECT id, room_id, msg_log_id, created_at FROM downstream_order_reviews WHERE id = :id"),
        {"id": review_id},
    ).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")

    room_id = row.get("room_id") or ""
    msg_log_id = row.get("msg_log_id")
    review_created_at = row.get("created_at")

    if not room_id:
        return {"messages": [], "trigger_msg_id": msg_log_id, "review_msg_ids": []}

    cols = (
        "id, msg_uid, instance_id, room_id, room_name, sender_id, sender_name, "
        "message_type, content_preview, ai_recognized, is_at_bot, oss_key, created_at"
    )

    if msg_log_id:
        before_rows = db.execute(text(
            f"SELECT {cols} FROM message_logs "
            "WHERE room_id = :room_id AND id < :msg_id "
            "ORDER BY id DESC LIMIT 20"
        ), {"room_id": room_id, "msg_id": msg_log_id}).mappings().all()

        trigger_row = db.execute(text(
            f"SELECT {cols} FROM message_logs WHERE id = :msg_id"
        ), {"msg_id": msg_log_id}).mappings().all()

        after_rows = db.execute(text(
            f"SELECT {cols} FROM message_logs "
            "WHERE room_id = :room_id AND id > :msg_id "
            "ORDER BY id ASC LIMIT 20"
        ), {"room_id": room_id, "msg_id": msg_log_id}).mappings().all()

        all_rows = list(reversed(before_rows)) + list(trigger_row) + list(after_rows)
    else:
        before_rows = db.execute(text(
            f"SELECT {cols} FROM message_logs "
            "WHERE room_id = :room_id AND created_at <= :ts "
            "ORDER BY created_at DESC, id DESC LIMIT 20"
        ), {"room_id": room_id, "ts": review_created_at}).mappings().all()

        after_rows = db.execute(text(
            f"SELECT {cols} FROM message_logs "
            "WHERE room_id = :room_id AND created_at > :ts "
            "ORDER BY created_at ASC, id ASC LIMIT 20"
        ), {"room_id": room_id, "ts": review_created_at}).mappings().all()

        all_rows = list(reversed(before_rows)) + list(after_rows)

    msg_ids = [r["id"] for r in all_rows]

    # 查询同群聊所有审核单的 msg_log_id，用于前端标注"已关联审核单"
    review_rows = db.execute(text(
        "SELECT msg_log_id FROM downstream_order_reviews "
        "WHERE room_id = :room_id AND msg_log_id IS NOT NULL"
    ), {"room_id": room_id}).mappings().all()
    review_msg_ids = [int(r["msg_log_id"]) for r in review_rows if r["msg_log_id"]]

    # 查询这些消息中图片/文件类型关联的审核记录的 attachment_base64
    image_map: dict[int, dict[str, str]] = {}
    if msg_ids:
        placeholders = ", ".join(f":mid{i}" for i in range(len(msg_ids)))
        id_params = {f"mid{i}": mid for i, mid in enumerate(msg_ids)}
        img_reviews = db.execute(text(
            f"SELECT msg_log_id, attachment_base64, attachment_mime "
            f"FROM downstream_order_reviews "
            f"WHERE msg_log_id IN ({placeholders}) AND attachment_base64 IS NOT NULL AND attachment_base64 != ''"
        ), id_params).mappings().all()
        for ir in img_reviews:
            image_map[int(ir["msg_log_id"])] = {
                "base64": ir["attachment_base64"],
                "mime": ir["attachment_mime"] or "image/png",
            }

    messages = []
    for r in all_rows:
        item = dict(r)
        for k, v in item.items():
            if hasattr(v, "strftime"):
                item[k] = v.strftime("%Y-%m-%d %H:%M:%S")
        # 优先使用 OSS 媒体 URL
        if item.get("oss_key"):
            item["media_url"] = f"/api/downstream-orders/media/{item['id']}"
        else:
            # 回退：使用审核记录中的 attachment_base64
            img_data = image_map.get(item["id"])
            if img_data:
                item["image_base64"] = img_data["base64"]
                item["image_mime"] = img_data["mime"]
        messages.append(item)

    return {"messages": messages, "trigger_msg_id": msg_log_id, "review_msg_ids": review_msg_ids}


def _load_customer(db: Session, customer_id: int) -> dict[str, Any]:
    customer = db.execute(
        text("SELECT id, customer_name, contact_person, phone, address, erp_customer_id FROM downstream_customers WHERE id = :id AND deleted_at IS NULL AND status = 1"),
        {"id": customer_id},
    ).mappings().first()
    if not customer:
        raise ValueError("客户不存在或已停用")
    return dict(customer)


def _review_order_data(row: dict[str, Any]) -> dict[str, Any]:
    manual_order = _json_loads(row.get("manual_order_json"), None)
    if manual_order:
        return manual_order
    parsed_order = _json_loads(row.get("parsed_order_json"), None)
    if parsed_order:
        return parsed_order
    raise ValueError("当前记录尚无可下单的解析结果")


def check_duplicate_order(db: Session, review_id: int, customer_id: int, order_data_override: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    """下单前检查该客户最近10张销售订单中是否存在整单完全相同的订单。
    整单匹配条件：明细行数相同，且每行的款号、颜色、尺码、数量完全一致。
    """
    row = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")
    customer = _load_customer(db, customer_id)
    erp_customer_id = customer.get("erp_customer_id") or ""
    if not erp_customer_id:
        return []

    if order_data_override:
        order_data = order_data_override
    else:
        order_data = _review_order_data(row)

    new_items = order_data.get("items", [])
    if not new_items:
        return []

    def _item_fingerprint(product_no: str, color: str, sizes: list) -> str:
        """生成单行明细的指纹：款号||颜色||尺码1:数量1|尺码2:数量2|..."""
        size_parts = sorted(
            f"{s.get('size', '')}:{s.get('qty', 0)}"
            for s in (sizes or []) if s.get("qty")
        )
        return f"{product_no}||{color}||{'|'.join(size_parts)}"

    def _order_fingerprint(item_fps: list[str]) -> str:
        """整单指纹：所有行指纹排序后拼接"""
        return "###".join(sorted(item_fps))

    new_fps = []
    for item in new_items:
        pno = item.get("product_no") or ""
        color = item.get("color") or ""
        sizes = item.get("sizes") or []
        new_fps.append(_item_fingerprint(pno, color, sizes))
    new_order_fp = _order_fingerprint(new_fps)

    recent_orders = db.execute(
        text(
            "SELECT order_no, order_date FROM erp_sales_orders "
            "WHERE customer_id = :cid "
            "ORDER BY order_date DESC, id DESC LIMIT 10"
        ),
        {"cid": erp_customer_id},
    ).mappings().all()

    if not recent_orders:
        return []

    order_nos = [o["order_no"] for o in recent_orders]
    order_dates = {o["order_no"]: o["order_date"] for o in recent_orders}
    placeholders = ", ".join(f":o{i}" for i in range(len(order_nos)))
    params = {f"o{i}": no for i, no in enumerate(order_nos)}
    existing_items = db.execute(
        text(
            f"SELECT order_no, product_no, color, total_qty, sizes_json "
            f"FROM erp_sales_order_items WHERE order_no IN ({placeholders})"
        ),
        params,
    ).mappings().all()

    orders_items: dict[str, list] = {}
    for ei in existing_items:
        orders_items.setdefault(ei["order_no"], []).append(ei)

    duplicates = []
    for ono, items_list in orders_items.items():
        if len(items_list) != len(new_items):
            continue
        fps = []
        for ei in items_list:
            sizes_raw = _json_loads(ei["sizes_json"], [])
            fps.append(_item_fingerprint(ei["product_no"] or "", ei["color"] or "", sizes_raw))
        if _order_fingerprint(fps) == new_order_fp:
            duplicates.append({
                "order_no": ono,
                "order_date": str(order_dates.get(ono, "")),
                "item_count": len(items_list),
            })
    return duplicates


async def approve_review(db: Session, review_id: int, customer_id: int, current_user: User, review_note: str = "") -> dict[str, Any]:
    import time as _t; _t0 = _t.time(); _logs = []
    def _log(msg): elapsed = round(_t.time() - _t0, 2); _logs.append(f"[{elapsed}s] {msg}"); _logger.info("[FLOW] %s", _logs[-1])
    _log("审核下单 开始")
    row = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")
    customer = _load_customer(db, customer_id)
    review_uid = row.get("review_uid") or ""
    erp_customer_id = customer.get("erp_customer_id") or ""
    existing_order = _check_review_uid_in_recent_orders(db, erp_customer_id, review_uid)
    if existing_order:
        db.execute(text(
            "UPDATE downstream_order_reviews SET review_status = 'exception', review_note = :note, "
            "reviewer_id = :rid, reviewer_name = :rname, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
        ), {"id": review_id, "note": f"重复下单拦截: 该审核单已在ERP订单 {existing_order} 中存在",
            "rid": current_user.id, "rname": current_user.real_name})
        db.commit()
        raise ValueError(f"该审核单已下过单，对应ERP订单号: {existing_order}，已标记为异常，请人工处理")
    order_data = _review_order_data(row)
    _inject_review_uid_to_remark(order_data, review_uid)
    _log("调用 ERP 创建订单+审核...")
    result = await erp_bridge.create_sales_order(order_data, customer)
    new_order_no = result.get("order_no") or ""
    _log(f"ERP 下单完成, 单号={new_order_no}")
    pnos = [item.get("product_no") for item in order_data.get("items", []) if item.get("product_no")]
    db.execute(
        text(
            "UPDATE downstream_order_reviews SET customer_id = :customer_id, customer_name = :customer_name, review_status = 'approved', erp_order_no = :erp_order_no, review_note = :review_note, reviewer_id = :reviewer_id, reviewer_name = :reviewer_name, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
        ),
        {
            "id": review_id,
            "customer_id": customer_id,
            "customer_name": customer.get("customer_name") or "",
            "erp_order_no": new_order_no,
            "review_note": review_note,
            "reviewer_id": current_user.id,
            "reviewer_name": current_user.real_name,
        },
    )
    db.commit()
    _log("DB更新完成")
    # 先同步订单明细（等待完成），再后台同步库存
    if new_order_no:
        try:
            from app.services.erp_sync import sync_single_order
            from app.services.erp_bridge import _erp_client
            if _erp_client:
                sync_result = await sync_single_order(_erp_client, new_order_no)
                _log(f"订单同步完成: {sync_result}")
        except Exception as sync_exc:
            _log(f"订单同步异常: {sync_exc}")
    _trigger_incremental_sync(order_no="", product_nos=pnos)
    _log("库存同步已触发")
    # 自动打印配货单
    auto_print_result = None
    if new_order_no:
        try:
            from app.services.printer_service import auto_print_picking_list
            auto_print_result = auto_print_picking_list(db, new_order_no)
            if auto_print_result:
                _log(f"自动打印: {'成功' if auto_print_result.get('printed') else '失败'}")
        except Exception as print_exc:
            _log(f"自动打印异常: {print_exc}")
    _log("全流程完成 ✅")
    resp = {**result, "review_status": "approved", "_debug_logs": _logs}
    if auto_print_result:
        resp["auto_print"] = auto_print_result
    return resp


async def replace_old_order(db: Session, review_id: int, customer_id: int, current_user: User, review_note: str = "") -> dict[str, Any]:
    import time as _t; _t0 = _t.time(); _logs = []
    def _log(msg): elapsed = round(_t.time() - _t0, 2); _logs.append(f"[{elapsed}s] {msg}"); _logger.info("[FLOW] %s", _logs[-1])
    _log("替换旧单 开始")
    row = db.execute(text("SELECT * FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")
    customer = _load_customer(db, customer_id)
    review_uid = row.get("review_uid") or ""
    erp_customer_id = customer.get("erp_customer_id") or ""
    if not erp_customer_id:
        raise ValueError("所选客户缺少 ERP 客户编号，无法查询未发货订单")
    existing_order = _check_review_uid_in_recent_orders(db, erp_customer_id, review_uid)
    if existing_order:
        db.execute(text(
            "UPDATE downstream_order_reviews SET review_status = 'exception', review_note = :note, "
            "reviewer_id = :rid, reviewer_name = :rname, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
        ), {"id": review_id, "note": f"重复下单拦截: 该审核单已在ERP订单 {existing_order} 中存在",
            "rid": current_user.id, "rname": current_user.real_name})
        db.commit()
        raise ValueError(f"该审核单已下过单，对应ERP订单号: {existing_order}，已标记为异常，请人工处理")
    order_data = _review_order_data(row)
    _inject_review_uid_to_remark(order_data, review_uid)
    # 查询该客户所有未发货订单（不限货号），时间范围放宽到 5 年
    begin_date = (datetime.now() - timedelta(days=1825)).strftime("%Y-%m-%d")
    end_date = datetime.now().strftime("%Y-%m-%d")
    _log("查询未发货报表...")
    unshipped_rows = await erp_bridge.query_unshipped(erp_customer_id, product_nos=None, dates=begin_date, datee=end_date)
    _log(f"查询完成, 找到 {len(unshipped_rows)} 行未发货")
    # 先取消该客户所有未发货订单行，再下新单
    _log("取消未发货订单...")
    cancel_result = await erp_bridge.cancel_unshipped([item["id"] for item in unshipped_rows])
    _log("取消完成, 创建新订单+审核...")
    create_result = await erp_bridge.create_sales_order(order_data, customer)
    new_order_no = create_result.get("order_no") or ""
    _log(f"新单创建完成, 单号={new_order_no}")
    pnos = [item.get("product_no") for item in order_data.get("items", []) if item.get("product_no")]
    replaced_orders = sorted({item.get("order_no") for item in unshipped_rows if item.get("order_no")})
    db.execute(
        text(
            "UPDATE downstream_order_reviews SET customer_id = :customer_id, customer_name = :customer_name, review_status = 'replaced', erp_order_no = :erp_order_no, replaced_order_no = :replaced_order_no, replace_source_ids = :replace_source_ids, review_note = :review_note, reviewer_id = :reviewer_id, reviewer_name = :reviewer_name, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
        ),
        {
            "id": review_id,
            "customer_id": customer_id,
            "customer_name": customer.get("customer_name") or "",
            "erp_order_no": new_order_no,
            "replaced_order_no": ",".join(replaced_orders),
            "replace_source_ids": _json_dumps(unshipped_rows),
            "review_note": review_note,
            "reviewer_id": current_user.id,
            "reviewer_name": current_user.real_name,
        },
    )
    db.commit()
    _log("DB更新完成")
    # 先同步订单明细（等待完成），再后台同步库存
    if new_order_no:
        try:
            from app.services.erp_sync import sync_single_order
            from app.services.erp_bridge import _erp_client
            if _erp_client:
                sync_result = await sync_single_order(_erp_client, new_order_no)
                _log(f"订单同步完成: {sync_result}")
        except Exception as sync_exc:
            _log(f"订单同步异常: {sync_exc}")
    _trigger_incremental_sync(order_no="", product_nos=pnos)
    _log("库存同步已触发")
    # 自动打印配货单
    auto_print_result = None
    if new_order_no:
        try:
            from app.services.printer_service import auto_print_picking_list
            auto_print_result = auto_print_picking_list(db, new_order_no)
            if auto_print_result:
                _log(f"自动打印: {'成功' if auto_print_result.get('printed') else '失败'}")
        except Exception as print_exc:
            _log(f"自动打印异常: {print_exc}")
    _log("全流程完成 ✅")
    resp = {**create_result, **cancel_result, "review_status": "replaced", "replaced_orders": replaced_orders, "_debug_logs": _logs}
    if auto_print_result:
        resp["auto_print"] = auto_print_result
    return resp


async def manual_order(db: Session, review_id: int, customer_id: int, order_data: dict[str, Any], current_user: User, review_note: str = "") -> dict[str, Any]:
    import time as _t; _t0 = _t.time(); _logs = []
    def _log(msg): elapsed = round(_t.time() - _t0, 2); _logs.append(f"[{elapsed}s] {msg}"); _logger.info("[FLOW] %s", _logs[-1])
    _log("手动录单 开始")
    row = db.execute(text("SELECT id, review_uid FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")
    customer = _load_customer(db, customer_id)
    review_uid = row.get("review_uid") or ""
    erp_customer_id = customer.get("erp_customer_id") or ""
    existing_order = _check_review_uid_in_recent_orders(db, erp_customer_id, review_uid)
    if existing_order:
        db.execute(text(
            "UPDATE downstream_order_reviews SET review_status = 'exception', review_note = :note, "
            "reviewer_id = :rid, reviewer_name = :rname, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
        ), {"id": review_id, "note": f"重复下单拦截: 该审核单已在ERP订单 {existing_order} 中存在",
            "rid": current_user.id, "rname": current_user.real_name})
        db.commit()
        raise ValueError(f"该审核单已下过单，对应ERP订单号: {existing_order}，已标记为异常，请人工处理")
    normalized = _normalize_order(order_data, customer.get("customer_name") or "")
    _inject_review_uid_to_remark(normalized, review_uid)
    _log("调用 ERP 创建订单+审核...")
    result = await erp_bridge.create_sales_order(normalized, customer)
    new_order_no = result.get("order_no") or ""
    _log(f"ERP 下单完成, 单号={new_order_no}")
    pnos = [item.get("product_no") for item in normalized.get("items", []) if item.get("product_no")]
    db.execute(
        text(
            "UPDATE downstream_order_reviews SET customer_id = :customer_id, customer_name = :customer_name, review_status = 'manual_ordered', manual_order_json = :manual_order_json, erp_order_no = :erp_order_no, review_note = :review_note, reviewer_id = :reviewer_id, reviewer_name = :reviewer_name, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
        ),
        {
            "id": review_id,
            "customer_id": customer_id,
            "customer_name": customer.get("customer_name") or "",
            "manual_order_json": _json_dumps(normalized),
            "erp_order_no": new_order_no,
            "review_note": review_note,
            "reviewer_id": current_user.id,
            "reviewer_name": current_user.real_name,
        },
    )
    db.commit()
    _log("DB更新完成")
    _trigger_incremental_sync(order_no=new_order_no, product_nos=pnos)
    _log("全流程完成 ✅")
    return {**result, "review_status": "manual_ordered", "_debug_logs": _logs}


def void_review(db: Session, review_id: int, current_user: User, customer_id: int | None = None, review_note: str = "") -> dict[str, Any]:
    row = db.execute(text("SELECT id, review_uid, customer_id FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("待审核记录不存在")
    review_uid = row.get("review_uid") or ""
    cid = customer_id or row.get("customer_id")
    if cid and review_uid:
        try:
            customer = _load_customer(db, cid)
            erp_customer_id = customer.get("erp_customer_id") or ""
            existing_order = _check_review_uid_in_recent_orders(db, erp_customer_id, review_uid)
            if existing_order:
                db.execute(text(
                    "UPDATE downstream_order_reviews SET review_status = 'exception', review_note = :note, "
                    "reviewer_id = :rid, reviewer_name = :rname, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
                ), {"id": review_id, "note": f"重复下单拦截: 该审核单已在ERP订单 {existing_order} 中存在",
                    "rid": current_user.id, "rname": current_user.real_name})
                db.commit()
                raise ValueError(f"该审核单已下过单，对应ERP订单号: {existing_order}，已标记为异常，请人工处理")
        except ValueError:
            raise
        except Exception:
            pass
    db.execute(
        text(
            "UPDATE downstream_order_reviews SET review_status = 'voided', review_note = :review_note, reviewer_id = :reviewer_id, reviewer_name = :reviewer_name, reviewed_at = NOW(), updated_at = NOW() WHERE id = :id"
        ),
        {
            "id": review_id,
            "review_note": review_note,
            "reviewer_id": current_user.id,
            "reviewer_name": current_user.real_name,
        },
    )
    db.commit()
    return {"review_status": "voided"}


def revert_to_pending(db: Session, review_id: int, current_user: User) -> dict[str, Any]:
    row = db.execute(text("SELECT id, review_status FROM downstream_order_reviews WHERE id = :id"), {"id": review_id}).mappings().first()
    if not row:
        raise ValueError("记录不存在")
    if row["review_status"] not in ("voided", "exception"):
        raise ValueError("只有废单或异常状态才能转为待审核")
    db.execute(
        text(
            "UPDATE downstream_order_reviews SET review_status = 'pending', reviewer_id = :reviewer_id, reviewer_name = :reviewer_name, reviewed_at = NULL, updated_at = NOW() WHERE id = :id"
        ),
        {
            "id": review_id,
            "reviewer_id": current_user.id,
            "reviewer_name": current_user.real_name,
        },
    )
    db.commit()
    return {"review_status": "pending"}

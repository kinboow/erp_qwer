"""
客户群 AI 对话服务 — 基于字节跳动豆包 API (OpenAI 兼容) 的 Function Calling 架构

每个被监听的客户群维护独立的对话上下文（以 room_id 为 key），
AI 通过工具自主完成：识别报货 → 查询商品 → 补全缺失信息 → 创建待审核订单。
"""

from __future__ import annotations

import base64
import io
import json
import logging
import secrets
import time
from datetime import datetime
from typing import Any, Optional

import httpx
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.config import settings
from app.database import SessionLocal

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
MAX_HISTORY_MESSAGES = 80   # 每个群最多保留的对话轮数
MAX_TOOL_ROUNDS = 8         # 单次消息最多连续 tool-call 轮数（防无限循环）
CHAT_API_TIMEOUT = 240      # AI API 超时（秒）

# ---------------------------------------------------------------------------
# 表创建（复用 downstream_support 统一管理）
# ---------------------------------------------------------------------------
def ensure_chat_table(db: Session) -> None:
    from app.services.downstream_support import ensure_downstream_support_tables
    ensure_downstream_support_tables(db)


def _load_ai_config(db: Session) -> dict[str, Any]:
    from app.services.ai_config import get_ai_config_for_parser

    try:
        return get_ai_config_for_parser(db)
    except Exception as exc:
        logger.warning("加载 AI 配置失败，回退 .env: %s", exc)
        return {
            "provider": "qwen",
            "base_url": settings.OPENAI_BASE_URL.rstrip("/"),
            "api_key": settings.OPENAI_API_KEY,
            "model": settings.OPENAI_MODEL,
            "vision_model": settings.OPENAI_VISION_MODEL,
            "temperature": 0.1,
            "enabled": True,
        }


async def _call_chat_api(cfg: dict[str, Any], messages: list[dict[str, Any]]) -> dict[str, Any]:
    from app.services.ai_config import log_ai_call

    if not cfg.get("enabled", True):
        raise RuntimeError("AI 已关闭")
    base_url = str(cfg.get("base_url") or "").rstrip("/")
    api_key = str(cfg.get("api_key") or "")
    model = str(cfg.get("model") or "")
    if not base_url or not api_key or not model:
        raise RuntimeError("AI 配置不完整")

    provider = str(cfg.get("provider") or "qwen")
    temperature = float(cfg.get("temperature") or 0.1)
    request_body: dict[str, Any] = {
        "model": model,
        "temperature": temperature,
        "messages": messages,
        "tools": TOOL_DEFINITIONS,
        "tool_choice": "auto",
        "max_tokens": 16384,
    }
    if provider == "bytedance":
        request_body["thinking"] = {"type": "enabled"}

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    req_summary_parts: list[str] = []
    for msg in messages[-20:]:
        role = str(msg.get("role") or "")
        content = msg.get("content")
        if isinstance(content, str):
            snippet = content[:300]
        elif isinstance(content, list):
            texts = [p.get("text", "") for p in content if isinstance(p, dict) and p.get("type") == "text"]
            suffix = " [+图片]" if any(isinstance(p, dict) and p.get("type") == "image_url" for p in content) else ""
            snippet = (" ".join(texts)[:300] + suffix).strip()
        else:
            snippet = ""
        req_summary_parts.append(f"[{role}] {snippet}"[:350])
    req_summary = "\n".join(req_summary_parts)[:4000]

    started = time.time()
    try:
        async with httpx.AsyncClient(timeout=CHAT_API_TIMEOUT) as client:
            response = await client.post(
                f"{base_url}/chat/completions",
                headers=headers,
                json=request_body,
            )
        if response.status_code >= 400:
            logger.error("客户群 AI API 请求失败 [%s %s]: status=%d body=%s", provider, model, response.status_code, response.text[:1000])
        response.raise_for_status()
        payload = response.json()
        duration_ms = int((time.time() - started) * 1000)
        usage = payload.get("usage") or {}
        choice = (payload.get("choices") or [{}])[0]
        message = choice.get("message") or {}
        resp_summary = ((message.get("content") or "") if isinstance(message.get("content"), str) else "")[:16000]
        tool_calls = message.get("tool_calls") or []
        if tool_calls:
            try:
                tc_names = ", ".join(str((tc.get("function") or {}).get("name") or "") for tc in tool_calls[:8])
                resp_summary = (resp_summary + f"\n[tool_calls] {tc_names}").strip()[:16000]
            except Exception:
                pass

        db = SessionLocal()
        try:
            log_ai_call(
                db,
                model=model,
                caller="customer_group_chat",
                prompt_tokens=usage.get("prompt_tokens", 0),
                completion_tokens=usage.get("completion_tokens", 0),
                total_tokens=usage.get("total_tokens", 0),
                duration_ms=duration_ms,
                status="success",
                request_summary=req_summary,
                response_summary=resp_summary,
            )
        finally:
            db.close()
        return payload
    except Exception as exc:
        duration_ms = int((time.time() - started) * 1000)
        db = SessionLocal()
        try:
            log_ai_call(
                db,
                model=model,
                caller="customer_group_chat",
                duration_ms=duration_ms,
                status="error",
                error_message=f"{type(exc).__name__}: {exc}"[:2000],
                request_summary=req_summary,
            )
        finally:
            db.close()
        raise


# ---------------------------------------------------------------------------
# System Prompt
# ---------------------------------------------------------------------------
CUSTOMER_GROUP_SYSTEM_PROMPT = """你是一个服装行业客户群的智能报货助手机器人，你在一个微信客户群中工作。

## 一、你的职责
1. 监听群里的聊天消息，识别客户的报货（下单）需求
2. 当客户发送报货信息时（包含货号、颜色、尺码、数量等），自动处理下单
3. 如果报货信息不完整（缺少货号、颜色、尺码或数量），主动@客户询问缺失信息
4. 确认信息完整后，调用工具创建待审核订单
5. 判断订单意图（新下单 / 替换旧单 / 追加补充）
6. 处理系统已经发出的“未发货确认”追问：客户自然回复后，你需要判断他是在说继续发货还是取消订单，并调用对应工具

## 二、消息格式
- 每条消息前面会标注消息ID和发送者姓名，格式为 "[消息ID:123][发送者姓名] 消息内容"
- 你能看到群里的客户消息，以及机器人账号自己发送的消息（非图片/Excel的普通文件仍不会送入你）
- 消息是实时推送给你的，每收到一条你就会处理一次
- 图片会以 base64 格式提供，表格文件会转为 Markdown 表格
- **撤回消息**：当客户撤回了一条消息，你会收到格式为 "[系统通知] XXX 撤回了一条消息。被撤回的原始内容：..." 的特殊消息
- **员工与客户区分**：系统会在提示词末尾提供本公司员工名单。员工发的消息不是报货，但你仍然可以看到员工消息作为上下文（例如员工在群里确认客户的订单、回复客户问题等）。
  只有非员工（客户）发的消息才可能是报货信息，员工发的消息永远不要当作报货处理。
- **当前实例账号消息**：你现在也会看到当前实例账号自己发送到群里的消息。这类消息会被明确标记为“当前实例账号发送”。它们只用于提供上下文，帮助你理解前后对话，绝不能当成客户报货、客户确认或员工接手的依据。

## 三、报货识别流程（核心，按顺序执行）

你需要完成以下步骤来处理每条可能的报货消息：

### 步骤1：判断消息是否包含可提取的报货数据（核心逻辑）
判断当前消息（文字/图片/表格）中是否包含可以解析为标准报货格式的数据（货号/款号、颜色、尺码、数量），分两种情况：

- 【包含报货数据 → 继续步骤2】消息中包含可以提取的报货数据（具体的货号/款号、产品名称/别名、颜色+尺码+数量等）。
  **不管客户是否明确说了"下单""报货"**，只要消息里有可识别的报货格式信息，就必须继续步骤2解析。
  后续由 check_existing_reviews 工具和上下文一起判断是否最终创建订单。
- 【不包含报货数据 → 不处理】日常闲聊、打招呼、问候、表情包、闲聊图片、讨论非订单话题、纯文字询问（没有具体款号/数量信息）

**关键原则：只要能提取出报货数据，就先提取，不要犹豫。后面的查重步骤会负责判断是否真的需要下单。**

### 步骤2：款号匹配（极其重要）
识别客户提到的产品，并匹配到本年产品目录中的标准货号：
- **优先调用 query_product_catalog 查询本年产品目录**，可直接不传 keyword，让工具返回完整的本年产品库列表，再从中匹配标准货号
- 如已基本确认款号，可再调用 query_product_details 核对颜色、尺码和别名信息
- 客户可能用货号、产品名称、别名/俗称来指代产品，你需要智能匹配
- **尾号简写匹配**：客户经常只说货号的后几位（尾号），你需要匹配到目录中尾号一致的完整款号。
  例如：目录中有"95862"，客户说"862"就是指它；目录中有"82842"，客户说"842"就是指它。
  如果尾号匹配到多个款号，需要询问客户确认具体是哪一个。
- **只有本年产品目录中存在的款号才能下单**，如果客户提到的款号不在目录中，告知客户该款号不存在或无法识别
- 如果客户提到的名称有多个可能的匹配，先询问客户确认

### 步骤3：颜色和尺码匹配
查询到款号后，用 query_product_details 获取该款号的可选颜色列表和可选尺码列表，然后做匹配：
- **颜色智能匹配规则（极其重要）：**
  - 手写简写/俗称/近义词必须智能对应到列表内最接近的标准名
  - 例如："米白"→"奶油白"、"黑"→"黑色"/"奢雅黑"、"花灰"→"花灰色"、"茶"→"茶色"、"藏青色"→"藏青"、"酱紫"→"酱紫色"
  - 判断标准：只要语义上指的是同一种颜色，就应该匹配到列表中对应的标准名称
  - 只有当客户提到的颜色与列表中所有颜色都完全无关时，才视为无效
- **尺码**必须从该款号的可选尺码列表中匹配

### 步骤4：检查报货信息完整性
完整的报货信息需要四个要素：**款号（货号）、颜色、尺码、数量**。
- 全部齐全 → 步骤5
- 缺少任意字段 → 用 send_group_message @客户，明确告知缺少什么（如"请问需要什么颜色和尺码？"）

### 步骤5：查重（极其重要，必须在创建订单前执行）
信息完整后，**必须先调用 check_existing_reviews 工具**，把解析好的标准报货数据传入，检查审核列表中是否已存在完全一致的报货记录。

#### 情况A：check_existing_reviews 返回 match=false（不一致 / 没找到重复）
说明这是一条全新的报货信息。结合上下文判断客户是否在下单：
- **确定是要报单** → 直接调用 create_order_review 创建订单，然后 send_group_message 回复"收到"
- **模棱两可，不太确定是不是要报单**（比如客户只发了图片/表格，但没说"下单""报货"） → **也默认去报单**，调用 create_order_review 创建订单，让人工审核去判断
- **明确不是报单**（如客户在问有没有货、能不能发、纯粹询价） → 不报单，忽略

#### 情况B：check_existing_reviews 返回 match=true（一致 / 已存在完全相同的报货）
说明审核列表中已有完全一样的报货记录（款号+颜色+尺码+数量全部一致）。此时结合上下文判断客户真实意图：
- **客户是在催发货/问物流**（如"这些什么时候发""能今天送吗""帮我催一下"） → **不报单、不调用创建订单工具**，用 send_group_message 回复客户，如"亲 这单已经在处理了哈~"
- **客户不是催发货，无法确定意图** → 用 send_group_message @客户 询问："亲 这一单和之前报的一样哦，是要再下一单还是在问之前那单呀？"
- **客户明确说要再下一单**（如"再来一单""这个再报一次"） → 调用 create_order_review 创建订单

### 步骤6：创建订单
只有经过步骤5判断后确认需要报单时，才调用 create_order_review 创建待审核订单，然后用 send_group_message 通知客户"收到"。
- **数据结构**：每个 item = 一个"款号+颜色"组合，sizes 数组包含该颜色下所有尺码和数量
- **合并规则**：同一款号同一颜色的不同尺码必须合并到同一个 item 的 sizes 数组中
  - 例："82761 白色 M 10件 L 5件" → 一个 item，sizes=[{size:"M", qty:10}, {size:"L", qty:5}]
  - 例："82761 白色 M 10件, 82761 黑色 L 5件" → 两个 item（不同颜色）
- **必须传 trigger_msg_log_id**：它必须是你真正据此识别并决定创建订单的那一条客户消息的消息ID。
  如果报货信息分散在多条消息里，传“最后那条让你确认可以下单/补齐关键信息”的消息ID，不要乱填。
- **数量必须是具体数字**，qty 必须大于 0，不能为空或为 0
- **"1"的歧义判断（重要）**：客户单独回复"1"或"好的""嗯"等，在上下文中往往是对上一条消息的确认/肯定，
  而不是下单1件。你必须结合上下文判断：如果前一条是疑问句（如"要不要XXX？""确定吗？"），
  回复"1"代表"确认/是的"，不应当作数量处理。只有在明确的报货语境中（如"82761 白色 M 1"）才把1视为数量
- **备注提取**：如果客户消息中包含与订单相关的备注说明（如发货方式、包装要求、加急、不要吊牌、和XXX一起发等），
  提取到 order_remark 字段。备注仅限与订单处理相关的内容，闲聊内容不算备注。
  如果某条备注仅针对特定款号/颜色，放到对应 item 的 remark 字段中。

## 四、名称映射与别名系统（重要）
- 产品目录中每个款号可能有**别名**（alias），客户可能用别名而非款号来指代产品
- 例如：客户说"弯刀裤"，实际对应货号"82761"；说"云朵裤"，对应"95890"
- 当客户表述模糊、使用俗称或你不确定时，优先调用 query_product_catalog，必要时可不传 keyword 直接拿完整目录和别名映射再匹配
- query_product_catalog 支持按别名搜索，query_product_details 也会返回别名信息
- 如果同一个名称可能匹配多个款号，必须先询问客户确认

## 五、图片和表格解析注意事项
1. **背面透字过滤**：如果图片中有纸张背面透过来的文字（颜色较浅、方向相反、镜像或模糊的印刷/手写痕迹），请完全忽略这些背面透字，只识别纸张正面清晰可见的内容
2. **手写内容**：手写内容模糊、涂改、无法辨认的货号/颜色/尺码/数量，**不要猜测**，
   必须调用 send_group_message @客户 在群里询问，明确指出哪些部分看不清（如"亲 图片上第2行的数量看不太清，麻烦确认一下~"）
3. **一次消息多商品**：一条消息/图片/表格中可能包含多个商品的报货信息，要全部识别，每个款号+颜色单独一条记录
4. **表格数据**：Excel/表格中的数据要完整提取，注意表头和数据行的对应关系
5. 只提取货号、颜色、尺码、数量、备注信息，**不需要识别客户名、联系人、下单日期等无关信息**

## 六、订单意图判断（极其重要，必须区分以下情况）

### 情况A：替换旧单（replace）— 取消 ERP 中未发货旧单，以新单为准
客户明确要**取消之前已下过的订单（已进入 ERP 系统的），全部作废，以这次新发的为准**，关键词：
- "替换""换单""重新报""之前那个不要了""把昨天的单改一下""这个替换之前的""作废之前的""全部换成这个"
→ 调用 create_order_review 时设置 order_intent="replace"
→ 审核人员审核时会先取消该客户所有未发货订单，再下新单

### 情况B：更正刚发的报货（写错了/发错了）— 作废审核单后重新下
客户只是**更正刚才发的报货内容**（还没审核下单的），关键词：
- "刚才那个写错了""发错了""重新发一下""那张不要了换这个""上面那个不对"
→ 先调用 void_recent_review 作废刚才的审核单，再用 create_order_review 创建新的（order_intent="new"）

### 情况C：修改旧单（modify）— 修改之前报过的单的部分内容
客户想**修改之前报过的单里面的部分信息**（如改颜色、改数量、去掉某一款、加几件），但不是全部取消重来，关键词：
- "之前那个XXX改成YYY""把82761白色改成黑色""上次报的那个数量改一下""帮我把M改成L""那个减2件""多加一个尺码"
→ 调用 create_modify_review 创建待修改审核单，说明要修改的内容

### 区分 replace 和 modify 的标准（极其重要）
- **replace（替换旧单）**：客户要把旧单**全部取消**，以新发的完整报货为准。新报货包含完整的货号+颜色+尺码+数量。
- **modify（修改旧单）**：客户**不取消旧单**，只是要**改其中某些信息**（改颜色、改尺码、改数量、增减某一款）。
- 如果你无法确定客户是要 replace 还是 modify，用 send_group_message @客户 询问：
  "亲 是要把之前的单全部取消重新报呢，还是只改其中某些内容呀？"

### 情况B 与 情况C 的区分
- 情况B 是**刚才几分钟内发错了**（还在待审核中），作废审核单重来
- 情况C 是**之前已经报过并可能已经审核下单了的**，客户回头想改部分内容

### 追加补充（append）
客户说"追加""再加""补几件""加一点" → order_intent="append"

如果没有明确的替换/追加/更正/修改意图，默认为 new。

### 情况D：客户撤回了报单消息
当你收到“[系统通知] XXX 撤回了一条消息”时，需要判断：
1. 查看被撤回的原始内容是否是报单/报货消息
2. 如果是报单消息，并且你之前已经为这条消息创建了待审核订单：
   - 系统会根据该消息对应的消息ID自动作废绑定的待审核订单
   - 你**不要**再调用 void_recent_review，也**不要**在群里发送任何回复，静默忽略即可
3. 如果被撤回的不是报单消息，或者你还没来得及处理该消息，直接忽略不处理

## 七、回复规范（极其重要）
- **绝大多数消息你都不需要回复，也不需要调用任何工具**。只有以下情况才行动：
  1. 识别到新的报货需求 → 查询产品目录/详情 → 创建订单
  2. 报货信息不完整，需要询问缺失字段 → send_group_message 询问
  3. 订单创建成功 → send_group_message 回复确认
  4. 客户撤回了报单消息 → 系统会按消息ID自动作废已绑定的待审核订单，你保持静默，不回复、不调用工具
  5. 客户要修改之前报过的单的部分内容 → create_modify_review 创建待修改审核单 → send_group_message 回复确认
  6. 不确定客户是要替换旧单还是修改旧单 → send_group_message 询问确认
  7. **客户是在回复“这张未发部分是继续还是取消”这类系统追问**：
     - 明确表示**继续发货** → 调用 `mark_followup_continue`，**不要调用 send_group_message，不要回复任何文本**
     - 明确表示**取消这张单/不要了/未发部分取消** → 调用 `create_cancel_unshipped_review`，**不要调用 send_group_message，不要回复任何文本**
     - 如果当前群里同时有多张待确认订单，而客户回复又无法判断对应哪一张 → 才允许调用 `send_group_message` 追问具体订单号
- **不需要回复的情况（直接忽略，不调用任何工具，不输出任何内容）：**
  - 日常闲聊、问候、表情包、与报货无关的对话
  - 客户询问有没有货、库存够不够、什么时候到货、能不能发货等咨询类问题 — 你不是客服，不负责回答这些问题，完全忽略
  - 客户回复了你之前询问的缺失信息（此时你应该继续处理订单，但不需要再次询问或重复确认）
  - 同一个报货任务的后续补充消息，如果前一条消息已经触发了处理流程，不要重复回复
  - 你已经对某条报货信息回复过确认，后续相同内容不要再次回复
  - **不确定是否为报单的消息**（参见步骤1 “不确定” 情况），在连续 3 条未确认之前不要主动询问
  - **任何超出你职责范围的问题**（如价格咨询、售后问题、物流查询、产品介绍等）— 你只负责报货接单，其他一律保持静默，不要调用 send_group_message，不要回复任何内容

## 八、语气和风格（必须遵守）
- **像真人客服一样说话**，不要像机器人。要口语化、自然、简短
- 适当使用语气词，如"哈""呢""啦""噢""好的""嗯"等
- 标点符号不要太正式，可以用~、！等，少用句号
- **订单创建成功后**只需要说类似"收到~""好的收到啦""知道了~"这样的短回复，**不要说"共X个订单"、不要列订单明细、不要总结**
- 询问缺失信息时语气轻松自然，例如"亲 颜色和尺码发一下呢~""这个要什么码呀"
- 不要用"您好""请问""感谢"等过于正式的词，用"亲""哈""呢"代替
- 回复尽量控制在一句话以内，能短则短
"""

# ---------------------------------------------------------------------------
# Tool Definitions (OpenAI function-calling format)
# ---------------------------------------------------------------------------
TOOL_DEFINITIONS: list[dict[str, Any]] = [
    {
        "type": "function",
        "function": {
            "name": "create_order_review",
            "description": (
                "当你确认客户要报货且信息完整（货号、颜色、尺码、数量齐全）时，调用此工具创建待审核订单。"
                "每个 item 是一个“款号+颜色”组合，sizes 数组包含该颜色下所有尺码和对应数量。"
                "同一款号同一颜色的不同尺码必须合并到同一个 item 的 sizes 数组中，不要拆分成多个 item。"
                "例如：客户说'82761 白色 M 10件 L 5件' → 一个 item: product_no=82761, color=白色, sizes=[{size:M, qty:10}, {size:L, qty:5}]"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "items": {
                        "type": "array",
                        "description": "订单项列表，每个 item 是一个“款号+颜色”组合",
                        "items": {
                            "type": "object",
                            "properties": {
                                "product_no": {"type": "string", "description": "货号/款号"},
                                "color": {"type": "string", "description": "颜色"},
                                "sizes": {
                                    "type": "array",
                                    "description": "该颜色下所有尺码和数量",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "size": {"type": "string", "description": "尺码"},
                                            "qty": {"type": "integer", "description": "数量，必须大于0"},
                                        },
                                        "required": ["size", "qty"],
                                    },
                                },
                                "remark": {"type": "string", "description": "备注（可选）"},
                            },
                            "required": ["product_no", "color", "sizes"],
                        },
                    },
                    "order_remark": {
                        "type": "string",
                        "description": "整单备注（可选），客户提到的与订单处理相关的备注说明，如发货方式、包装要求、加急等。仅提取与订单相关的内容，闲聊不算。",
                    },
                    "order_intent": {
                        "type": "string",
                        "enum": ["new", "replace", "append"],
                        "description": "订单意图: new=新下单, replace=替换旧单, append=追加补充。默认 new。",
                    },
                    "trigger_msg_log_id": {
                        "type": "integer",
                        "description": "真正触发本次下单识别的客户消息ID。必须填写当前对话里那条实际报货/补齐关键信息的消息ID。",
                    },
                },
                "required": ["items", "trigger_msg_log_id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "query_product_catalog",
            "description": (
                "查询本年款产品目录（仅包含本年度在售产品），获取可用的货号、产品名称、别名映射。"
                "客户提到的货号、产品名、别名/俗称都可以作为关键词搜索。"
                "只有本年目录中存在的款号才能创建订单，不在目录中的款号不能下单。"
                "当客户使用模糊叫法、俗称、尾号或你暂时无法确认具体款号时，应优先直接调用本工具且不传 keyword，"
                "先获取完整的本年产品库列表及别名映射，再进行匹配。可传入关键词搜索，也可留空获取全部产品。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "keyword": {
                        "type": "string",
                        "description": "搜索关键词（货号、产品名称或别名），留空则返回全部目录",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "query_product_details",
            "description": (
                "查询指定货号的详细信息，包括该款号的可选颜色列表和可选尺码列表。"
                "在创建订单前必须先查询此工具，确认客户提供的颜色和尺码在该款号的可选范围内。"
                "也支持通过别名查询，如传入别名会自动映射到对应的标准货号。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "product_no": {
                        "type": "string",
                        "description": "货号/款号",
                    },
                },
                "required": ["product_no"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "void_recent_review",
            "description": (
                "作废/撤销当前群最近创建的一条待审核订单。"
                "当客户说刚才那个写错了、发错了、重新发、那张不要了，指的是更正刚刚提交的报货内容（而不是替换 ERP 中已有的旧订单），"
                "你应该先调用此工具作废之前那条审核单，再创建新的正确订单。"
                "可选传入 review_id 指定作废哪条，不传则自动作废当前群最近的一条 pending 审核单。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "review_id": {
                        "type": "integer",
                        "description": "要作废的审核单 ID（可选，不传则自动找最近一条）",
                    },
                    "reason": {
                        "type": "string",
                        "description": "作废原因，如'客户重新发送了正确的报货信息'",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_modify_review",
            "description": (
                "当客户要修改之前已经报过的单（不是全部取消重来，而是改其中部分内容）时调用此工具。"
                "创建一条'待修改'类型的审核单，人工审核员会根据修改说明去 ERP 中手动修改对应订单。"
                "modify_description 要清晰描述客户要修改什么，比如'82761白色改成黑色'、'M码减2件'等。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "modify_description": {
                        "type": "string",
                        "description": "客户要修改的具体内容描述，如'82761 白色改成黑色，M码数量从10改成8'",
                    },
                    "original_items": {
                        "type": "array",
                        "description": "修改涉及的原始款号信息（可选，帮助审核员定位）",
                        "items": {
                            "type": "object",
                            "properties": {
                                "product_no": {"type": "string", "description": "涉及的货号"},
                                "color": {"type": "string", "description": "涉及的颜色（可选）"},
                            },
                        },
                    },
                },
                "required": ["modify_description"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "send_group_message",
            "description": (
                "向当前客户群发送消息。当需要向客户确认信息、询问缺失字段、"
                "回复收到等时调用。设置 at_sender=true 可以 @触发消息的发送者。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "message": {
                        "type": "string",
                        "description": "要发送的消息文本",
                    },
                    "at_sender": {
                        "type": "boolean",
                        "description": "是否@触发消息的发送者",
                        "default": False,
                    },
                },
                "required": ["message"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "mark_followup_continue",
            "description": (
                "当客户是在回答系统之前发出的未发货确认追问，并明确表示这张订单未发部分要继续发货时调用。"
                "调用后系统会把该订单标记为继续发货，并在后续第5天若仍未发完时再次追问。"
                "调用此工具后不要再发送群消息。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "order_no": {
                        "type": "string",
                        "description": "客户确认继续发货的订单号。必须是当前群待确认列表中的订单号。"
                    }
                },
                "required": ["order_no"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_cancel_unshipped_review",
            "description": (
                "当客户是在回答系统之前发出的未发货确认追问，并明确表示这张订单未发部分要取消时调用。"
                "调用后系统会创建一条人工审核单，审核员点击审核即取消该订单未发货部分。"
                "调用此工具后不要再发送群消息。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "order_no": {
                        "type": "string",
                        "description": "客户要求取消未发部分的订单号。必须填写。"
                    },
                    "reason": {
                        "type": "string",
                        "description": "客户取消原因或你对客户原话的简要总结，可选。"
                    }
                },
                "required": ["order_no"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "check_existing_reviews",
            "description": (
                "查询当前群的审核列表中，是否已经存在与传入报货信息完全一致的待审核/已审核订单。"
                "完全一致指：每一个款号+颜色+尺码+数量都一模一样。"
                "你必须在调用 create_order_review 之前，先调用此工具检查是否重复。"
                "根据返回结果决定后续动作（见系统提示词中的报货查重流程）。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "items": {
                        "type": "array",
                        "description": "要检查的报货信息列表，格式与 create_order_review 的 items 完全一致",
                        "items": {
                            "type": "object",
                            "properties": {
                                "product_no": {"type": "string", "description": "货号/款号"},
                                "color": {"type": "string", "description": "颜色"},
                                "sizes": {
                                    "type": "array",
                                    "description": "该颜色下所有尺码和数量",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "size": {"type": "string", "description": "尺码"},
                                            "qty": {"type": "integer", "description": "数量"},
                                        },
                                        "required": ["size", "qty"],
                                    },
                                },
                            },
                            "required": ["product_no", "color", "sizes"],
                        },
                    },
                },
                "required": ["items"],
            },
        },
    },
]


# ---------------------------------------------------------------------------
# Tool Execution
# ---------------------------------------------------------------------------
def _tool_create_order_review(
    db: Session,
    room_id: str,
    sender_id: str,
    sender_name: str,
    customer: Optional[dict[str, Any]],
    instance_id: str,
    args: dict[str, Any],
) -> str:
    """工具: 创建待审核订单"""
    from app.services.downstream_support import ensure_downstream_support_tables
    ensure_downstream_support_tables(db)

    raw_items = args.get("items") or []
    if not raw_items:
        return json.dumps({"ok": False, "error": "items 为空，无法创建订单"}, ensure_ascii=False)
    trigger_msg_log_id_raw = args.get("trigger_msg_log_id")
    try:
        trigger_msg_log_id = int(trigger_msg_log_id_raw or 0)
    except Exception:
        trigger_msg_log_id = 0
    if trigger_msg_log_id <= 0:
        return json.dumps({"ok": False, "error": "trigger_msg_log_id 缺失或无效，必须标注真正触发下单的消息ID"}, ensure_ascii=False)

    trigger_msg = db.execute(
        text(
            "SELECT id, room_id, sender_id, sender_name, message_type, message_server_id "
            "FROM message_logs WHERE id = :id LIMIT 1"
        ),
        {"id": trigger_msg_log_id},
    ).mappings().first()
    if not trigger_msg:
        return json.dumps({"ok": False, "error": f"trigger_msg_log_id={trigger_msg_log_id} 对应的消息不存在"}, ensure_ascii=False)
    if str(trigger_msg.get("room_id") or "").strip() != str(room_id or "").strip():
        return json.dumps({"ok": False, "error": f"trigger_msg_log_id={trigger_msg_log_id} 不属于当前群，不能用于本次下单"}, ensure_ascii=False)

    # 校验并规范化 items：过滤掉 qty<=0 的 size，过滤掉 sizes 为空的 item
    items: list[dict[str, Any]] = []
    for it in raw_items:
        pno = str(it.get("product_no") or "").strip()
        color = str(it.get("color") or "").strip()
        remark = str(it.get("remark") or "").strip()
        sizes = []
        for s in (it.get("sizes") or []):
            size_name = str(s.get("size") or "").strip()
            qty = int(s.get("qty") or 0)
            if size_name and qty > 0:
                sizes.append({"size": size_name, "qty": qty})
        if pno and color and sizes:
            items.append({"product_no": pno, "color": color, "sizes": sizes, "remark": remark})

    if not items:
        return json.dumps({"ok": False, "error": "items 中无有效的尺码数量数据（qty 必须 > 0）"}, ensure_ascii=False)

    order_intent = args.get("order_intent", "new")
    order_remark = str(args.get("order_remark") or "").strip()
    review_uid = f"RV{datetime.now().strftime('%y%m%d')}{secrets.token_hex(2).upper()}"

    # 直接创建审核记录，parse_status=success（AI 已完成解析）
    order_json: dict[str, Any] = {"items": items}
    if order_remark:
        order_json["remark"] = order_remark
    parsed_order_json = json.dumps(order_json, ensure_ascii=False)
    content_summary = "; ".join(
        f"{it['product_no']} {it['color']} " + "/".join(f"{s['size']}x{s['qty']}" for s in it['sizes'])
        for it in items
    )

    result = db.execute(
        text(
            "INSERT INTO downstream_order_reviews ("
            "review_uid, source_type, instance_id, room_id, sender_id, sender_name, "
            "message_type, content_text, parse_status, review_status, "
            "customer_id, customer_name, parsed_order_json, order_intent, operator_name, msg_log_id"
            ") VALUES ("
            ":review_uid, 'ai_conversation', :instance_id, :room_id, :sender_id, :sender_name, "
            "'text', :content_text, 'success', 'pending', "
            ":customer_id, :customer_name, :parsed_order_json, :order_intent, '机器人', :msg_log_id"
            ")"
        ),
        {
            "review_uid": review_uid,
            "instance_id": instance_id or None,
            "room_id": room_id,
            "sender_id": str(trigger_msg.get("sender_id") or sender_id or "").strip(),
            "sender_name": str(trigger_msg.get("sender_name") or sender_name or "").strip(),
            "content_text": content_summary,
            "customer_id": customer["id"] if customer else None,
            "customer_name": customer.get("customer_name", "") if customer else "",
            "parsed_order_json": parsed_order_json,
            "order_intent": order_intent,
            "msg_log_id": trigger_msg_log_id,
        },
    )
    review_id = result.lastrowid
    db.commit()

    # SSE 通知前端
    try:
        from app.services.review_events import notify_review_change
        notify_review_change("new_review", {"review_id": review_id})
    except Exception:
        pass

    return json.dumps({
        "ok": True,
        "review_id": review_id,
        "review_uid": review_uid,
        "items_count": len(items),
        "order_intent": order_intent,
        "trigger_msg_log_id": trigger_msg_log_id,
        "trigger_message_server_id": str(trigger_msg.get("message_server_id") or ""),
    }, ensure_ascii=False)


def void_reviews_for_recalled_message(db: Session, room_id: str, msg_log_id: int, reason: str) -> dict[str, Any]:
    from app.services.downstream_support import ensure_downstream_support_tables
    ensure_downstream_support_tables(db)

    if msg_log_id <= 0:
        return {"ok": False, "count": 0, "review_ids": []}

    rows = db.execute(
        text(
            "SELECT id, review_uid FROM downstream_order_reviews "
            "WHERE room_id = :room_id AND msg_log_id = :msg_log_id AND review_status = 'pending'"
        ),
        {"room_id": room_id, "msg_log_id": msg_log_id},
    ).mappings().all()
    if not rows:
        return {"ok": True, "count": 0, "review_ids": []}

    review_ids = [int(row["id"]) for row in rows if row.get("id") is not None]
    db.execute(
        text(
            "UPDATE downstream_order_reviews SET review_status = 'voided', review_note = :note, "
            "operator_name = '机器人', updated_at = NOW() "
            "WHERE room_id = :room_id AND msg_log_id = :msg_log_id AND review_status = 'pending'"
        ),
        {"note": reason, "room_id": room_id, "msg_log_id": msg_log_id},
    )
    db.commit()

    try:
        from app.services.review_events import notify_review_change
        for review_id in review_ids:
            notify_review_change("new_review", {"review_id": review_id})
    except Exception:
        pass
    return {"ok": True, "count": len(review_ids), "review_ids": review_ids}


def _tool_void_recent_review(
    db: Session,
    room_id: str,
    args: dict[str, Any],
) -> str:
    """工具: 作废当前群最近的一条待审核订单"""
    review_id = args.get("review_id")
    reason = args.get("reason") or "AI 自动作废（客户更正报货内容）"

    if review_id:
        row = db.execute(
            text("SELECT id, review_uid, review_status FROM downstream_order_reviews WHERE id = :id AND room_id = :room_id"),
            {"id": review_id, "room_id": room_id},
        ).mappings().first()
    else:
        row = db.execute(
            text(
                "SELECT id, review_uid, review_status FROM downstream_order_reviews "
                "WHERE room_id = :room_id AND review_status = 'pending' "
                "ORDER BY id DESC LIMIT 1"
            ),
            {"room_id": room_id},
        ).mappings().first()

    if not row:
        return json.dumps({"ok": False, "error": "未找到可作废的待审核订单"}, ensure_ascii=False)

    if row["review_status"] != "pending":
        return json.dumps({"ok": False, "error": f"该审核单状态为 {row['review_status']}，只能作废 pending 状态的单"}, ensure_ascii=False)

    db.execute(
        text(
            "UPDATE downstream_order_reviews SET review_status = 'voided', review_note = :note, operator_name = '机器人', updated_at = NOW() WHERE id = :id"
        ),
        {"id": row["id"], "note": reason},
    )
    db.commit()

    # SSE 通知前端
    try:
        from app.services.review_events import notify_review_change
        notify_review_change("new_review", {"review_id": row["id"]})
    except Exception:
        pass

    return json.dumps({
        "ok": True,
        "voided_review_id": row["id"],
        "voided_review_uid": row["review_uid"],
    }, ensure_ascii=False)


def _tool_create_modify_review(
    db: Session,
    room_id: str,
    sender_id: str,
    sender_name: str,
    customer: Optional[dict[str, Any]],
    instance_id: str,
    args: dict[str, Any],
) -> str:
    """工具: 创建待修改审核单（客户要修改之前报过的单的部分内容）"""
    from app.services.downstream_support import ensure_downstream_support_tables
    ensure_downstream_support_tables(db)

    modify_desc = str(args.get("modify_description") or "").strip()
    if not modify_desc:
        return json.dumps({"ok": False, "error": "modify_description 不能为空"}, ensure_ascii=False)

    original_items = args.get("original_items") or []
    review_uid = f"RV{datetime.now().strftime('%y%m%d')}{secrets.token_hex(2).upper()}"

    # 将修改信息存入 parsed_order_json
    order_json: dict[str, Any] = {
        "modify_description": modify_desc,
        "original_items": original_items,
    }
    parsed_order_json = json.dumps(order_json, ensure_ascii=False)

    result = db.execute(
        text(
            "INSERT INTO downstream_order_reviews ("
            "review_uid, source_type, instance_id, room_id, sender_id, sender_name, "
            "message_type, content_text, parse_status, review_status, review_type, "
            "customer_id, customer_name, parsed_order_json, order_intent, operator_name"
            ") VALUES ("
            ":review_uid, 'ai_conversation', :instance_id, :room_id, :sender_id, :sender_name, "
            "'text', :content_text, 'success', 'pending', 'modify', "
            ":customer_id, :customer_name, :parsed_order_json, 'modify', '机器人'"
            ")"
        ),
        {
            "review_uid": review_uid,
            "instance_id": instance_id or None,
            "room_id": room_id,
            "sender_id": sender_id,
            "sender_name": sender_name,
            "content_text": f"[待修改] {modify_desc}",
            "customer_id": customer["id"] if customer else None,
            "customer_name": customer.get("customer_name", "") if customer else "",
            "parsed_order_json": parsed_order_json,
        },
    )
    review_id = result.lastrowid
    db.commit()

    # SSE 通知前端
    try:
        from app.services.review_events import notify_review_change
        notify_review_change("new_review", {"review_id": review_id})
    except Exception:
        pass

    return json.dumps({
        "ok": True,
        "review_id": review_id,
        "review_uid": review_uid,
        "type": "modify",
        "modify_description": modify_desc,
    }, ensure_ascii=False)


def _tool_mark_followup_continue(
    db: Session,
    room_id: str,
    args: dict[str, Any],
) -> str:
    from app.services.printer_service import mark_followup_continue_decision

    order_no = str(args.get("order_no") or "").strip()
    if not order_no:
        return json.dumps({"ok": False, "error": "order_no 不能为空"}, ensure_ascii=False)
    try:
        result = mark_followup_continue_decision(db, room_id, order_no)
    except Exception as exc:
        return json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False)
    return json.dumps(result, ensure_ascii=False)


def _tool_create_cancel_unshipped_review(
    db: Session,
    room_id: str,
    sender_id: str,
    sender_name: str,
    customer: Optional[dict[str, Any]],
    instance_id: str,
    args: dict[str, Any],
) -> str:
    from app.services.printer_service import get_room_pending_followups, mark_followup_cancel_review_created

    order_no = str(args.get("order_no") or "").strip()
    reason = str(args.get("reason") or "").strip()
    if not order_no:
        return json.dumps({"ok": False, "error": "order_no 不能为空"}, ensure_ascii=False)

    pending_followups = get_room_pending_followups(db, room_id)
    matched = next((item for item in pending_followups if str(item.get("order_no") or "").strip() == order_no), None)
    if not matched:
        return json.dumps({"ok": False, "error": f"当前群未找到订单 {order_no} 的待确认记录"}, ensure_ascii=False)

    review_uid = f"RV{datetime.now().strftime('%y%m%d')}{secrets.token_hex(2).upper()}"
    item_summary = matched.get("item_summary") or []
    order_json = {
        "cancel_order_no": order_no,
        "decision": "cancel_unshipped",
        "followup_stage": matched.get("current_stage") or "third_day",
        "item_summary": item_summary,
        "reason": reason,
    }
    parsed_order_json = json.dumps(order_json, ensure_ascii=False)
    content_text = f"[取消未发货] 订单 {order_no} 客户确认取消未发部分"
    if reason:
        content_text = f"{content_text}：{reason}"

    result = db.execute(
        text(
            "INSERT INTO downstream_order_reviews ("
            "review_uid, source_type, instance_id, room_id, sender_id, sender_name, "
            "message_type, content_text, parse_status, review_status, review_type, "
            "customer_id, customer_name, parsed_order_json, order_intent, operator_name, replaced_order_no"
            ") VALUES ("
            ":review_uid, 'ai_conversation', :instance_id, :room_id, :sender_id, :sender_name, "
            "'text', :content_text, 'success', 'pending', 'cancel_unshipped', "
            ":customer_id, :customer_name, :parsed_order_json, 'replace', '机器人', :replaced_order_no"
            ")"
        ),
        {
            "review_uid": review_uid,
            "instance_id": instance_id or None,
            "room_id": room_id,
            "sender_id": sender_id,
            "sender_name": sender_name,
            "content_text": content_text,
            "customer_id": customer["id"] if customer else None,
            "customer_name": customer.get("customer_name", "") if customer else "",
            "parsed_order_json": parsed_order_json,
            "replaced_order_no": order_no,
        },
    )
    review_id = int(result.lastrowid or 0)
    db.commit()
    if review_id:
        mark_followup_cancel_review_created(db, room_id, order_no, review_id)

    try:
        from app.services.review_events import notify_review_change
        notify_review_change("new_review", {"review_id": review_id})
    except Exception:
        pass

    return json.dumps({
        "ok": True,
        "review_id": review_id,
        "review_uid": review_uid,
        "review_type": "cancel_unshipped",
        "order_no": order_no,
    }, ensure_ascii=False)


def _tool_check_existing_reviews(
    db: Session,
    room_id: str,
    args: dict[str, Any],
) -> str:
    """工具: 查询审核列表中是否存在与当前报货完全一致的记录。

    硬代码精确比对：款号、颜色、尺码、数量每一项都一模一样才算"一致"。
    """
    check_items = args.get("items") or []
    if not check_items:
        return json.dumps({"ok": False, "error": "items 为空"}, ensure_ascii=False)

    # 将输入的 items 标准化为可比较的 fingerprint
    def _fingerprint(items_list: list[dict[str, Any]]) -> set[tuple]:
        fp: set[tuple] = set()
        for it in items_list:
            pno = str(it.get("product_no") or "").strip()
            color = str(it.get("color") or "").strip()
            for s in (it.get("sizes") or []):
                size = str(s.get("size") or "").strip()
                qty = int(s.get("qty") or 0)
                if pno and color and size and qty > 0:
                    fp.add((pno, color, size, qty))
        return fp

    input_fp = _fingerprint(check_items)
    if not input_fp:
        return json.dumps({"ok": False, "error": "无有效的报货数据"}, ensure_ascii=False)

    # 查同群最近 30 天内 pending / approved 的审核单
    rows = db.execute(
        text(
            "SELECT id, review_uid, review_status, parsed_order_json, sender_name, created_at "
            "FROM downstream_order_reviews "
            "WHERE room_id = :room_id "
            "  AND review_status IN ('pending', 'approved') "
            "  AND review_type = 'normal' "
            "  AND created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY) "
            "ORDER BY id DESC LIMIT 50"
        ),
        {"room_id": room_id},
    ).mappings().all()

    for row in rows:
        try:
            parsed = json.loads(row["parsed_order_json"] or "{}")
        except Exception:
            continue
        existing_items = parsed.get("items") or []
        if not existing_items:
            continue
        existing_fp = _fingerprint(existing_items)
        if existing_fp == input_fp:
            return json.dumps({
                "ok": True,
                "match": True,
                "matched_review_id": row["id"],
                "matched_review_uid": row["review_uid"],
                "matched_review_status": row["review_status"],
                "matched_sender_name": row.get("sender_name") or "",
                "matched_created_at": str(row.get("created_at") or ""),
                "message": "审核列表中已存在完全一致的报货记录",
            }, ensure_ascii=False)

    return json.dumps({
        "ok": True,
        "match": False,
        "message": "审核列表中未找到完全一致的报货记录",
    }, ensure_ascii=False)


def _tool_query_product_catalog(db: Session, args: dict[str, Any]) -> str:
    """工具: 查询产品目录"""
    from app.services.downstream_orders import query_current_year_catalog
    catalog = query_current_year_catalog(db)

    keyword = (args.get("keyword") or "").strip().lower()
    if keyword:
        raw_tokens = [t.strip().lower() for t in keyword.replace("，", " ").replace(",", " ").split() if t.strip()]
        tokens = raw_tokens or [keyword]
        filtered = []
        for item in catalog:
            pno = (item.get("product_no") or "").lower()
            pname = (item.get("product_name") or "").lower()
            aliases = [a.lower() for a in (item.get("aliases") or [])]
            haystacks = [pno, pname, *aliases]
            matched = False
            for token in tokens:
                if token in pno or token in pname or any(token in a for a in aliases):
                    matched = True
                    break
                if token.isdigit() and len(token) >= 3:
                    if pno.endswith(token) or any(a.endswith(token) for a in aliases):
                        matched = True
                        break
            if matched or any(token == h for token in tokens for h in haystacks if h):
                filtered.append(item)
        catalog = filtered

    # 限制返回数量避免 token 过长
    if len(catalog) > 50:
        return json.dumps({
            "total": len(catalog),
            "showing": 50,
            "hint": "结果过多，请用更精确的关键词搜索",
            "products": catalog[:50],
        }, ensure_ascii=False)

    return json.dumps({
        "total": len(catalog),
        "products": catalog,
    }, ensure_ascii=False)


def _tool_query_product_details(db: Session, args: dict[str, Any]) -> str:
    """工具: 查询指定货号详情"""
    from app.services.downstream_orders import query_current_year_catalog
    catalog = query_current_year_catalog(db)

    target_pno = (args.get("product_no") or "").strip()
    if not target_pno:
        return json.dumps({"ok": False, "error": "未指定货号"}, ensure_ascii=False)
    normalized_target = target_pno.lower()

    # 精确匹配
    for item in catalog:
        if str(item["product_no"] or "").strip().lower() == normalized_target:
            colors = [c.strip() for c in (item.get("color") or "").split(",") if c.strip()]
            sizes = [s.strip() for s in (item.get("spec") or "").split(",") if s.strip()]
            return json.dumps({
                "ok": True,
                "product_no": item["product_no"],
                "product_name": item.get("product_name", ""),
                "aliases": item.get("aliases", []),
                "available_colors": colors,
                "available_sizes": sizes,
            }, ensure_ascii=False)

    # 别名匹配
    for item in catalog:
        aliases = [str(a).strip() for a in (item.get("aliases") or []) if str(a).strip()]
        aliases_lower = [a.lower() for a in aliases]
        if normalized_target in aliases_lower:
            colors = [c.strip() for c in (item.get("color") or "").split(",") if c.strip()]
            sizes = [s.strip() for s in (item.get("spec") or "").split(",") if s.strip()]
            return json.dumps({
                "ok": True,
                "product_no": item["product_no"],
                "product_name": item.get("product_name", ""),
                "aliases": aliases,
                "matched_via_alias": target_pno,
                "available_colors": colors,
                "available_sizes": sizes,
            }, ensure_ascii=False)

    # 尾号匹配
    if normalized_target.isdigit() and len(normalized_target) >= 3:
        candidates: list[dict[str, Any]] = []
        for item in catalog:
            pno = str(item.get("product_no") or "").strip()
            aliases = [str(a).strip() for a in (item.get("aliases") or []) if str(a).strip()]
            if pno.endswith(normalized_target) or any(a.lower().endswith(normalized_target) for a in aliases):
                candidates.append({
                    "product_no": pno,
                    "product_name": item.get("product_name", ""),
                    "aliases": aliases,
                })
        if len(candidates) == 1:
            item = candidates[0]
            full = next((c for c in catalog if str(c.get("product_no") or "").strip() == item["product_no"]), None)
            colors = [c.strip() for c in ((full or {}).get("color") or "").split(",") if c.strip()]
            sizes = [s.strip() for s in ((full or {}).get("spec") or "").split(",") if s.strip()]
            return json.dumps({
                "ok": True,
                "product_no": item["product_no"],
                "product_name": item.get("product_name", ""),
                "aliases": item.get("aliases", []),
                "matched_via_tail": target_pno,
                "available_colors": colors,
                "available_sizes": sizes,
            }, ensure_ascii=False)
        if len(candidates) > 1:
            return json.dumps({
                "ok": False,
                "error": f"尾号 {target_pno} 匹配到多个款号，请让客户确认",
                "candidates": candidates,
            }, ensure_ascii=False)

    return json.dumps({"ok": False, "error": f"未找到货号 {target_pno}"}, ensure_ascii=False)


async def _tool_send_group_message(
    room_id: str,
    sender_id: str,
    instance_id: str,
    args: dict[str, Any],
) -> str:
    """工具: 向群里发送消息"""
    from app.services.wechat_reply import send_room_at
    db = SessionLocal()
    try:
        message = args.get("message", "")
        at_sender = args.get("at_sender", False)
        at_list = [sender_id] if at_sender and sender_id else None

        result = await send_room_at(
            db, room_id, message,
            at_list=at_list,
            instance_id=int(instance_id) if instance_id and instance_id.isdigit() else None,
        )
        return json.dumps({"ok": result.get("ok", False)}, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 对话历史管理
# ---------------------------------------------------------------------------
def _save_message(db: Session, room_id: str, role: str, content: Any = None,
                  name: str = None, tool_calls: Any = None, tool_call_id: str = None) -> None:
    """保存一条对话消息到 DB"""
    ensure_chat_table(db)
    content_str = json.dumps(content, ensure_ascii=False) if isinstance(content, (list, dict)) else content
    tool_calls_str = json.dumps(tool_calls, ensure_ascii=False) if tool_calls else None
    db.execute(
        text(
            "INSERT INTO ai_chat_messages (room_id, role, content, name, tool_calls, tool_call_id) "
            "VALUES (:room_id, :role, :content, :name, :tool_calls, :tool_call_id)"
        ),
        {
            "room_id": room_id,
            "role": role,
            "content": content_str,
            "name": name,
            "tool_calls": tool_calls_str,
            "tool_call_id": tool_call_id,
        },
    )
    db.commit()


def _load_history(db: Session, room_id: str, limit: int = MAX_HISTORY_MESSAGES) -> list[dict[str, Any]]:
    """加载该群的最近 N 条对话历史，组装为 OpenAI messages 格式"""
    ensure_chat_table(db)
    rows = db.execute(
        text(
            "SELECT role, content, name, tool_calls, tool_call_id "
            "FROM ai_chat_messages "
            "WHERE room_id = :room_id "
            "ORDER BY id DESC LIMIT :limit"
        ),
        {"room_id": room_id, "limit": limit},
    ).mappings().all()

    messages: list[dict[str, Any]] = []
    for r in reversed(rows):  # 恢复时间正序
        msg: dict[str, Any] = {"role": r["role"]}
        content = r["content"]

        # 尝试解析 JSON content（multimodal）
        if content and content.startswith("["):
            try:
                msg["content"] = json.loads(content)
            except Exception:
                msg["content"] = content
        else:
            msg["content"] = content

        if r["name"]:
            msg["name"] = r["name"]
        if r["tool_calls"]:
            try:
                msg["tool_calls"] = json.loads(r["tool_calls"]) if isinstance(r["tool_calls"], str) else r["tool_calls"]
            except Exception:
                pass
        if r["tool_call_id"]:
            msg["tool_call_id"] = r["tool_call_id"]

        messages.append(msg)

    return messages


def _trim_history(db: Session, room_id: str, keep: int = MAX_HISTORY_MESSAGES) -> None:
    """清理超出保留数量的旧消息"""
    try:
        db.execute(
            text(
                "DELETE FROM ai_chat_messages "
                "WHERE room_id = :room_id AND id NOT IN ("
                "  SELECT id FROM (SELECT id FROM ai_chat_messages WHERE room_id = :room_id ORDER BY id DESC LIMIT :keep) t"
                ")"
            ),
            {"room_id": room_id, "keep": keep},
        )
        db.commit()
    except Exception as exc:
        logger.debug("清理对话历史异常: %s", exc)


# ---------------------------------------------------------------------------
# CDN 附件下载
# ---------------------------------------------------------------------------
async def _cdn_download_once(
    runtime: dict[str, Any],
    cdn_params: dict[str, Any],
    save_path: "Path",
) -> str:
    """单次 CDN 下载尝试，成功返回 base64，失败抛异常。"""
    from pathlib import Path as _Path

    mode = cdn_params.get("mode", "wx_download")
    api_route = f"cdn/{mode}"
    body = {k: v for k, v in cdn_params.items() if k != "mode"}
    body["save_path"] = str(save_path)

    headers: dict[str, str] = {}
    if runtime.get("api_key"):
        headers["X-API-Key"] = runtime["api_key"]

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            f"{runtime['api_base_url']}/api/{runtime['wxid']}/{api_route}",
            json=body,
            headers=headers,
        )
        resp.raise_for_status()
        resp_data = resp.json()

    if isinstance(resp_data, dict) and resp_data.get("code") not in (0, None):
        raise RuntimeError(resp_data.get("msg") or "CDN API 返回错误")

    actual_path = save_path
    if not actual_path.is_file():
        data_body = resp_data.get("data") if isinstance(resp_data.get("data"), dict) else {}
        for key in ("save_path", "path", "file_path"):
            possible = str(data_body.get(key) or "").strip()
            if possible and _Path(possible).is_file():
                actual_path = _Path(possible)
                break
    if not actual_path.is_file():
        raise RuntimeError("下载后文件不存在")

    file_bytes = actual_path.read_bytes()
    try:
        actual_path.unlink(missing_ok=True)
    except Exception:
        pass
    return base64.b64encode(file_bytes).decode("ascii")


async def _download_attachment_from_cdn(
    db: Session,
    payload: dict[str, Any],
    instance_id: str,
    message_type: str,
    *,
    room_id: str = "",
    room_name: str = "",
    file_name: str = "",
) -> str:
    """从企微 CDN 下载附件，带重试与备用模式，返回 base64。

    重试策略：
      1) 首选模式 × 1 次
      2) 同模式重试 × 1 次
      3) 备用模式 × 2 次（如果 payload 中有两种 CDN 参数）
    全部失败后向通知群发送告警。
    """
    from pathlib import Path
    from app.services.media_archive import _extract_all_cdn_candidates, _resolve_runtime, _guess_extension
    import asyncio as _asyncio

    candidates = _extract_all_cdn_candidates(payload)
    if not candidates:
        logger.debug("CDN下载: 无CDN参数，跳过")
        return ""

    runtime = _resolve_runtime(db, instance_id)
    if not runtime.get("api_base_url") or not runtime.get("wxid"):
        logger.debug("CDN下载: 缺少运行时配置")
        return ""

    ext = _guess_extension(file_name or "", message_type)
    download_dir = Path(__file__).resolve().parents[2] / "temp" / "ai_chat_attachments"
    download_dir.mkdir(parents=True, exist_ok=True)

    # 构建尝试队列：首选模式 × 2 + 备用模式 × 2
    primary = candidates[0]
    alternate = candidates[1] if len(candidates) > 1 else None
    attempts: list[tuple[str, dict[str, Any]]] = [
        (f"首选({primary.get('mode')})", primary),
        (f"重试({primary.get('mode')})", primary),
    ]
    if alternate:
        attempts.append((f"备用({alternate.get('mode')})-1", alternate))
        attempts.append((f"备用({alternate.get('mode')})-2", alternate))

    last_err = ""
    for label, params in attempts:
        save_path = download_dir / f"chat_{secrets.token_hex(4)}{ext}"
        try:
            result_b64 = await _cdn_download_once(runtime, params, save_path)
            logger.info("CDN下载成功 [%s]: %d bytes", label, len(result_b64) * 3 // 4)
            return result_b64
        except Exception as exc:
            last_err = str(exc)
            logger.warning("CDN下载失败 [%s]: %s", label, exc)
            await _asyncio.sleep(1)

    # 全部失败 → 通知群告警
    display_name = file_name or ("图片" if message_type in ("image", "img", "picture") else "文件")
    display_room = room_name or room_id or "未知群"
    alert_msg = f"⚠️ 客户群文件下载失败\n群聊：{display_room}\n文件：{display_name}\n原因：{last_err}\n已重试 {len(attempts)} 次均失败，请人工处理。"
    try:
        await _notify_download_failure(db, alert_msg)
    except Exception as notify_exc:
        logger.warning("下载失败通知发送异常: %s", notify_exc)

    return ""


async def _notify_download_failure(db: Session, message: str) -> None:
    """向所有通知群发送下载失败告警。"""
    from app.services.wechat_reply import send_room_at
    try:
        rows = db.execute(
            text("SELECT room_id FROM downstream_customer_wechat_rooms WHERE room_type = 'notification'")
        ).mappings().all()
    except Exception:
        rows = []
    for row in rows:
        try:
            await send_room_at(db, row["room_id"], message)
        except Exception as exc:
            logger.warning("通知群推送失败 room=%s: %s", row["room_id"], exc)


async def _detect_image_rotation_angle(
    db: Session,
    attachment_base64: str,
    *,
    file_name: str = "",
    content_text: str = "",
) -> int:
    from app.services.ai_order_parser import ai_order_parser

    if not attachment_base64:
        return 0

    lower_name = (file_name or "").lower()
    mime_type = "image/png" if lower_name.endswith(".png") else "image/jpeg"
    prompt_text = (
        "判断这张图片为了变成正向阅读，还需要旋转多少度。"
        "正向阅读指中文、表格、订单内容应正常直立阅读。"
        "只允许返回 JSON：{\"angle\": 0|90|180|-90, \"confidence\": \"high|medium|low\", \"reason\": \"...\"}。"
        "其中 angle 的含义是：需要对当前图片进行顺时针旋转多少度；-90 表示逆时针 90 度；"
        "如果图片本来就是正的或无法判断，请返回 0。"
    )
    if content_text:
        prompt_text += f" 附加上下文：{content_text[:120]}"

    try:
        cfg = ai_order_parser._load_config(db)
        ai_order_parser._ensure_enabled(cfg)
        result = await ai_order_parser._chat(
            cfg["vision_model"],
            [
                {"role": "system", "content": "你是图片方向判断助手，只能输出 JSON，不要输出任何额外文字。"},
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt_text},
                        {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{attachment_base64}"}},
                    ],
                },
            ],
            db=db,
            caller="image_orientation_detector",
        )
        raw_angle = result.get("angle", 0)
        try:
            angle = int(raw_angle)
        except Exception:
            angle = 0
        if angle == 270:
            angle = -90
        elif angle == -270:
            angle = 90
        if angle not in {0, 90, 180, -90}:
            angle = 0
        logger.info(
            "图片方向检测: angle=%s confidence=%s reason=%s file=%s",
            angle,
            result.get("confidence", ""),
            str(result.get("reason", ""))[:200],
            file_name or "",
        )
        return angle
    except Exception as exc:
        logger.warning("图片方向检测失败，回退原图: file=%s err=%s", file_name or "", exc)
        return 0


async def _normalize_image_orientation(
    db: Session,
    attachment_base64: str,
    *,
    file_name: str = "",
    content_text: str = "",
) -> tuple[str, int]:
    from app.services.ai_order_parser import rotate_images_in_messages

    angle = await _detect_image_rotation_angle(
        db,
        attachment_base64,
        file_name=file_name,
        content_text=content_text,
    )
    if angle == 0:
        return attachment_base64, 0

    lower_name = (file_name or "").lower()
    mime_type = "image/png" if lower_name.endswith(".png") else "image/jpeg"
    rotated = rotate_images_in_messages(
        [{"type": "image", "base64": attachment_base64, "mime": mime_type}],
        angle,
    )
    rotated_b64 = ((rotated[0] or {}).get("base64") if rotated else "") or attachment_base64
    return rotated_b64, angle


# ---------------------------------------------------------------------------
# Excel → Markdown 转换 (复用)
# ---------------------------------------------------------------------------
def _excel_to_markdown(attachment_base64: str) -> str:
    """将 Excel 附件转为 Markdown 表格"""
    if not attachment_base64 or load_workbook is None:
        return ""
    try:
        binary = base64.b64decode(attachment_base64)
        workbook = load_workbook(io.BytesIO(binary), data_only=True)
    except Exception as exc:
        return f"Excel 解析失败: {exc}"
    parts: list[str] = []
    for sheet in workbook.worksheets[:3]:
        parts.append(f"## Sheet: {sheet.title}\n")
        rows_data: list[list[str]] = []
        for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
            row_values = ["" if c is None else str(c).strip() for c in row[:20]]
            if any(row_values):
                rows_data.append(row_values)
        if not rows_data:
            parts.append("（空表）\n")
            continue
        max_cols = max(len(r) for r in rows_data)
        for r in rows_data:
            while len(r) < max_cols:
                r.append("")
        header = rows_data[0]
        parts.append("| " + " | ".join(h or " " for h in header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        for data_row in rows_data[1:]:
            parts.append("| " + " | ".join(data_row) + " |")
        parts.append("")
    return "\n".join(parts).strip()


# ---------------------------------------------------------------------------
# 消息批次窗口：同一个群 35 秒内的消息合并后统一送 AI
# ---------------------------------------------------------------------------
import asyncio as _asyncio

BATCH_WINDOW_SECONDS = 22  # 批次等待窗口
BACKLOG_WINDOW_SECONDS = 90

# room_id → {"task": asyncio.Task, "trigger_ts": float, "count": int,
#             "customer": dict|None, "instance_id": str, "senders": dict}
_room_batches: dict[str, dict[str, Any]] = {}
_wechat_reconnect_backlog_until: float = 0.0


def mark_wechat_reconnect_backlog_window(seconds: int = BACKLOG_WINDOW_SECONDS) -> None:
    global _wechat_reconnect_backlog_until
    _wechat_reconnect_backlog_until = max(_wechat_reconnect_backlog_until, time.time() + max(1, seconds))
    logger.info("客户群 AI 已开启企微恢复历史消息窗口: %ss", seconds)


def _remaining_wechat_reconnect_backlog_seconds() -> int:
    remaining = int(_wechat_reconnect_backlog_until - time.time())
    return remaining if remaining > 0 else 0


# ---------------------------------------------------------------------------
# 动态 System Prompt（注入员工名单）
# ---------------------------------------------------------------------------
_employee_cache: dict[str, Any] = {"names": [], "pairs": [], "ts": 0.0}


def _get_employee_identities(db: Session) -> list[dict[str, str]]:
    import time as _t
    now = _t.time()
    if now - _employee_cache["ts"] > 300 or not _employee_cache["pairs"]:
        try:
            rows = db.execute(
                text("SELECT wxid, nickname FROM wechat_employee_accounts ORDER BY id")
            ).mappings().all()
            pairs: list[dict[str, str]] = []
            names: list[str] = []
            for row in rows:
                wxid = str(row.get("wxid") or "").strip()
                nickname = str(row.get("nickname") or "").strip()
                if wxid or nickname:
                    pairs.append({"wxid": wxid, "nickname": nickname})
                if nickname:
                    names.append(nickname)
            _employee_cache["pairs"] = pairs
            _employee_cache["names"] = names
        except Exception:
            _employee_cache["pairs"] = []
            _employee_cache["names"] = []
        _employee_cache["ts"] = now
    return list(_employee_cache["pairs"])


def _build_system_prompt(db: Session) -> str:
    """在静态 prompt 基础上追加本公司员工名单，缓存 5 分钟。"""
    _get_employee_identities(db)
    names = _employee_cache["names"]
    if not names:
        return CUSTOMER_GROUP_SYSTEM_PROMPT

    employee_section = "\n\n## 附录：本公司员工名单\n以下是本公司员工在群里的昵称，这些人发的消息不是报货，仅作为上下文参考：\n"
    employee_section += "、".join(names)
    return CUSTOMER_GROUP_SYSTEM_PROMPT + employee_section


def _build_batch_runtime_context(
    db: Session,
    room_id: str,
    *,
    batch_started_at: float,
    is_history_backlog: bool,
) -> dict[str, Any]:
    from app.services.printer_service import get_room_pending_followups

    employee_identities = _get_employee_identities(db)
    employee_names = {item.get("nickname") or "" for item in employee_identities if item.get("nickname")}
    employee_wxids = {item.get("wxid") or "" for item in employee_identities if item.get("wxid")}
    rows = db.execute(
        text(
            "SELECT sender_id, sender_name, content_preview "
            "FROM message_logs WHERE room_id = :room_id AND created_at >= :started_at ORDER BY id ASC LIMIT 200"
        ),
        {"room_id": room_id, "started_at": datetime.fromtimestamp(batch_started_at)},
    ).mappings().all()
    employee_names_seen: set[str] = set()
    employee_messages: list[str] = []
    employee_replied = False
    for row in rows:
        sender_id = str(row.get("sender_id") or "").strip()
        sender_name = str(row.get("sender_name") or "").strip()
        if sender_id not in employee_wxids and sender_name not in employee_names:
            continue
        employee_replied = True
        if sender_name:
            employee_names_seen.add(sender_name)
        preview = str(row.get("content_preview") or "").strip()
        if preview:
            employee_messages.append(f"- [{sender_name or sender_id}] {preview[:120]}")
    return {
        "is_history_backlog": is_history_backlog,
        "employee_replied": employee_replied,
        "employee_names_seen": sorted(employee_names_seen),
        "employee_messages": employee_messages[:8],
        "pending_followups": get_room_pending_followups(db, room_id),
    }


def _build_runtime_instruction(context: dict[str, Any]) -> str:
    lines: list[str] = []
    if context.get("is_history_backlog"):
        lines.append("## 当前批次说明")
        lines.append("本次输入的是企业微信掉线恢复后补推的历史消息，不是实时新消息。")
        lines.append("你必须优先判断这批历史消息里，是否已经有本公司员工在群里接手、答复或处理客户需求。")
        lines.append("如果员工已经接手处理，则你必须保持静默：不要回复，不要调用工具，不要重复创建订单。")
    employee_names_seen = context.get("employee_names_seen") or []
    if employee_names_seen:
        lines.append(f"本批次中检测到的员工发送者：{'、'.join(employee_names_seen)}")
    employee_messages = context.get("employee_messages") or []
    if employee_messages:
        lines.append("员工在本批次中的消息摘要：")
        lines.extend(employee_messages)
    if context.get("employee_replied"):
        lines.append("系统检测到本批次历史消息中已有员工发言。你必须优先判断员工是否已经接手处理；若已接手，则保持静默，不要继续自动处理。")
    if context.get("last_sender_is_bot"):
        lines.append("## 当前批次特殊说明")
        lines.append("当前批次最后一条消息是当前实例账号自己发送到群里的消息。")
        lines.append("这类消息只用于补充上下文，通常不代表客户有了新的需求。")
        lines.append("如果当前批次里没有新的客户消息提供额外信息，你通常应该保持静默，不要回复，不要调用工具。")
    pending_followups = context.get("pending_followups") or []
    if pending_followups:
        lines.append("## 当前群待确认的未发货订单")
        lines.append("以下订单是系统之前已经发过追问、正在等待客户确认‘继续发货 / 取消’的订单。")
        lines.append("如果客户当前消息是在回答这些追问，你必须优先处理这些跟进，不要把它误判成新的报货。")
        for item in pending_followups[:10]:
            summary = item.get("item_summary") or []
            summary_text = "；".join(str(x) for x in summary[:3]) if summary else "-"
            lines.append(
                f"- 订单号：{item.get('order_no') or '-'}；阶段：{item.get('current_stage') or '-'}；已追问次数：{item.get('ask_count') or 0}；最近追问日期：{item.get('last_asked_date') or '-'}；未发摘要：{summary_text}"
            )
        lines.append("处理规则：若客户明确说继续，就调用 mark_followup_continue；若明确说取消，就调用 create_cancel_unshipped_review；这两种情况都不要回复群消息。")
    return "\n".join(lines).strip()


# ---------------------------------------------------------------------------
# 核心：处理客户群消息
# ---------------------------------------------------------------------------
async def process_customer_group_message(
    db: Session,
    *,
    room_id: str,
    sender_id: str,
    sender_name: str,
    message_type: str,
    content_text: str,
    attachment_base64: str = "",
    file_name: str = "",
    customer: Optional[dict[str, Any]] = None,
    instance_id: str = "",
    bot_wxid: str = "",
    payload: Optional[dict[str, Any]] = None,
    log_id: Optional[int] = None,
    is_history_backlog: bool = False,
) -> dict[str, Any]:
    """处理来自客户群的一条消息：先保存到对话历史，再通过 35 秒批次窗口统一送 AI。

    返回: {"ok": True/False, "batched": bool, ...}
    """
    ensure_chat_table(db)
    is_bot_account_message = bool(bot_wxid and sender_id and str(sender_id).strip() == str(bot_wxid).strip())

    # ---- 0. 图片/文件缺失 base64 时尝试从 CDN 下载 ----
    is_media = message_type in ("image", "img", "picture", "file")
    if is_media and not attachment_base64 and payload:
        room_name_display = (customer or {}).get("room_name") or room_id
        logger.info("图片/文件缺 base64，尝试 CDN 下载: room=%s type=%s", room_id, message_type)
        attachment_base64 = await _download_attachment_from_cdn(
            db, payload, instance_id, message_type,
            room_id=room_id, room_name=room_name_display, file_name=file_name,
        )

    if message_type in ("image", "img", "picture") and attachment_base64:
        attachment_base64, rotated_angle = await _normalize_image_orientation(
            db,
            attachment_base64,
            file_name=file_name,
            content_text=content_text,
        )
        if rotated_angle:
            logger.info("客户群图片已自动转正: room=%s sender=%s angle=%d", room_id, sender_name or sender_id, rotated_angle)

    # ---- 1. 构建 user message content ----
    sender_label = f"当前实例账号发送:{sender_name or sender_id}" if is_bot_account_message else (sender_name or sender_id)
    prefix = f"[消息ID:{int(log_id)}][{sender_label}] " if log_id else f"[{sender_label}] "
    user_content: Any  # str 或 list (multimodal)

    if message_type in ("image", "img", "picture") and attachment_base64:
        user_content = [
            {"type": "text", "text": prefix + (content_text or "（发送了一张图片）")},
            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{attachment_base64}"}},
        ]
    elif message_type == "file" and file_name.lower().endswith((".xlsx", ".xls")) and attachment_base64:
        md_table = _excel_to_markdown(attachment_base64)
        if md_table:
            user_content = prefix + f"（发送了表格文件 {file_name}）\n\n{md_table}"
        else:
            user_content = prefix + f"（发送了文件 {file_name}，无法解析）"
    else:
        user_content = prefix + (content_text or "")

    # ---- 2. 立即保存 user message 到对话历史 ----
    _save_message(db, room_id, "user", content=user_content, name=sender_label)
    logger.info("消息已入库: room=%s sender=%s type=%s bot_self=%s", room_id, sender_name, message_type, is_bot_account_message)

    # ---- 2.5 熔断期间缓冲消息信息（消息仍入库，只是不送AI） ----
    from app.services.ai_circuit_breaker import is_tripped as _cb_is_tripped, buffer_message as _cb_buffer
    if _cb_is_tripped():
        _cb_buffer({
            "room_id": room_id,
            "sender_id": sender_id,
            "sender_name": sender_name,
            "customer_name": (customer or {}).get("customer_name", ""),
            "content_preview": (content_text or "")[:200],
            "message_type": message_type,
        })

    # ---- 3. 批次窗口：同群消息合并后统一送 AI ----
    is_media_msg = message_type in ("image", "img", "picture", "file")
    batch = _room_batches.get(room_id)

    if batch:
        # 如果当前批次已包含图片/文件，且本条也是图片/文件 → 拆分批次
        # 立即结束旧批次送 AI，新消息另起新批次
        if is_media_msg and batch.get("media_count", 0) >= 1:
            logger.info("批次拆分: room=%s 旧批次含 %d 张图片/文件, 新图片/文件到达 → 立即提交旧批次",
                        room_id, batch.get("media_count", 0))
            old_task = batch.get("task")
            if old_task and not old_task.done():
                old_task.cancel()
            # 立即异步触发旧批次的 AI 处理
            _room_batches.pop(room_id, None)
            _asyncio.create_task(_batch_delayed_ai_call_immediate(room_id, batch))
            # 本条消息 fall-through 到下方创建新批次
            batch = None

    if batch:
        # 批次已在进行中 — 追加计数，刷新发送者信息，跳过 AI
        batch["count"] += 1
        if is_media_msg:
            batch["media_count"] = batch.get("media_count", 0) + 1
        batch["senders"][sender_id] = sender_name
        batch["last_sender_is_bot"] = is_bot_account_message
        batch["is_history_backlog"] = batch.get("is_history_backlog", False) or is_history_backlog
        if customer:
            batch["customer"] = customer
        if instance_id:
            batch["instance_id"] = instance_id
        if is_history_backlog:
            batch["window_seconds"] = max(batch.get("window_seconds", BATCH_WINDOW_SECONDS), BACKLOG_WINDOW_SECONDS)
        logger.info("消息加入批次窗口: room=%s count=%d media=%d 剩余 %.0fs",
                    room_id, batch["count"], batch.get("media_count", 0),
                    max(0, batch["trigger_ts"] + batch.get("window_seconds", BATCH_WINDOW_SECONDS) - time.time()))
        return {"ok": True, "batched": True, "batch_count": batch["count"]}

    # 无活跃批次 — 此消息为触发消息，创建新批次
    effective_window_seconds = BACKLOG_WINDOW_SECONDS if is_history_backlog else BATCH_WINDOW_SECONDS
    batch_info: dict[str, Any] = {
        "trigger_ts": time.time(),
        "count": 1,
        "media_count": 1 if is_media_msg else 0,
        "customer": dict(customer) if customer else None,
        "instance_id": instance_id or "",
        "senders": {sender_id: sender_name},
        "last_sender_is_bot": is_bot_account_message,
        "is_history_backlog": is_history_backlog,
        "window_seconds": effective_window_seconds,
        "task": None,
    }
    _room_batches[room_id] = batch_info

    # 启动延迟任务
    task = _asyncio.create_task(_batch_delayed_ai_call(room_id, batch_info))
    batch_info["task"] = task

    logger.info("批次窗口开启: room=%s 等待 %ds history_backlog=%s", room_id, effective_window_seconds, is_history_backlog)
    return {"ok": True, "batched": True, "batch_trigger": True}


async def _batch_delayed_ai_call_immediate(room_id: str, batch_info: dict[str, Any]) -> None:
    """立即触发旧批次的 AI 处理（用于批次拆分场景：新图片/文件到达时提交旧批次）。

    与 _batch_delayed_ai_call 共享后半段逻辑，只是跳过 sleep 等待。
    等 3 秒让旧批次中最后一条消息的 CDN 下载 / 入库完成。
    """
    await _asyncio.sleep(3)
    msg_count = batch_info.get("count", 0)
    logger.info("批次拆分提交: room=%s 立即处理旧批次 (%d 条消息)", room_id, msg_count)
    await _batch_run_ai(room_id, batch_info)


async def _batch_delayed_ai_call(room_id: str, batch_info: dict[str, Any]) -> None:
    """等待批次窗口结束后，加载完整历史并统一调用 AI。"""
    try:
        await _asyncio.sleep(batch_info.get("window_seconds", BATCH_WINDOW_SECONDS))
    except _asyncio.CancelledError:
        _room_batches.pop(room_id, None)
        logger.info("批次任务被取消: room=%s", room_id)
        return

    logger.info("批次窗口到期: room=%s 共 %d 条消息", room_id, batch_info.get("count", 0))
    await _batch_run_ai(room_id, batch_info)


async def _batch_run_ai(room_id: str, batch_info: dict[str, Any]) -> None:
    """批次 AI 处理的公共逻辑（被 _batch_delayed_ai_call 和 _batch_delayed_ai_call_immediate 复用）。"""
    msg_count = batch_info.get("count", 0)
    customer = batch_info.get("customer")
    instance_id = batch_info.get("instance_id", "")
    senders = batch_info.get("senders", {})
    last_sender_is_bot = bool(batch_info.get("last_sender_is_bot", False))
    is_history_backlog = bool(batch_info.get("is_history_backlog", False))
    batch_started_at = float(batch_info.get("trigger_ts") or time.time())
    # 取最后一个发送者作为 sender_id / sender_name（工具调用可能需要）
    last_sender_id = list(senders.keys())[-1] if senders else ""
    last_sender_name = senders.get(last_sender_id, "")

    logger.info("开始 AI 处理: room=%s 共 %d 条消息 history_backlog=%s", room_id, msg_count, is_history_backlog)

    # 熔断检查：如果 AI 已暂停，缓冲消息而不调用
    from app.services.ai_circuit_breaker import is_tripped, buffer_message
    if is_tripped():
        logger.warning("[AI熔断] 批次窗口到期但 AI 已暂停，缓冲消息: room=%s count=%d", room_id, msg_count)
        buffer_message({
            "room_id": room_id,
            "sender_id": last_sender_id,
            "sender_name": last_sender_name,
            "customer_name": (customer or {}).get("customer_name", ""),
            "content_preview": f"[批次 {msg_count} 条消息]",
            "message_type": "batch",
        })
        return

    # 获取同群锁 + 全局信号量，与 _send_msg_to_ai 共享，避免并发冲突
    from app.services.wechat_runtime_compat import _get_room_lock, _get_ai_semaphore
    room_lock = _get_room_lock(room_id)
    async with room_lock:
        if _room_batches.get(room_id) is batch_info:
            _room_batches.pop(room_id, None)
        async with _get_ai_semaphore():
            db = SessionLocal()
            try:
                await _run_ai_on_history(
                    db,
                    room_id=room_id,
                    sender_id=last_sender_id,
                    sender_name=last_sender_name,
                    customer=customer,
                    instance_id=instance_id,
                    last_sender_is_bot=last_sender_is_bot,
                    is_history_backlog=is_history_backlog,
                    batch_started_at=batch_started_at,
                )
            except Exception as exc:
                logger.error("批次 AI 处理异常: room=%s err=%s", room_id, exc, exc_info=True)
            finally:
                db.close()


async def _run_ai_on_history(
    db: Session,
    *,
    room_id: str,
    sender_id: str,
    sender_name: str,
    customer: Optional[dict[str, Any]],
    instance_id: str,
    last_sender_is_bot: bool = False,
    is_history_backlog: bool = False,
    batch_started_at: float | None = None,
) -> dict[str, Any]:
    """加载对话历史并调用 AI（多轮 tool-call），用于批次窗口到期后的统一处理。"""

    # ---- 1. 加载对话历史 ----
    history = _load_history(db, room_id)
    runtime_context = _build_batch_runtime_context(
        db,
        room_id,
        batch_started_at=batch_started_at or time.time(),
        is_history_backlog=is_history_backlog,
    )
    runtime_context["last_sender_is_bot"] = last_sender_is_bot
    runtime_instruction = _build_runtime_instruction(runtime_context)
    suppress_actions = bool(runtime_context.get("is_history_backlog") and runtime_context.get("employee_replied"))

    # ---- 2. 组装 API messages (system + history) ----
    system_prompt = _build_system_prompt(db)
    api_messages: list[dict[str, Any]] = [
        {"role": "system", "content": system_prompt},
    ]
    if runtime_instruction:
        api_messages.append({"role": "system", "content": runtime_instruction})
    for i, msg in enumerate(history):
        if isinstance(msg.get("content"), list) and i < len(history) - 5:
            text_parts = [p.get("text", "") for p in msg["content"] if p.get("type") == "text"]
            api_messages.append({"role": msg["role"], "content": " ".join(text_parts) + " [图片已省略]"})
        else:
            api_messages.append(msg)

    # ---- 3. 调用 AI (可能多轮 tool call) ----
    from app.services.ai_circuit_breaker import is_tripped, record_success, record_error
    if is_tripped():
        logger.warning("AI 已熔断，跳过对话: room=%s", room_id)
        return {"ok": False, "reason": "ai_tripped"}

    cfg = _load_ai_config(db)
    if not cfg.get("enabled"):
        logger.info("AI 已关闭，跳过对话: room=%s", room_id)
        return {"ok": False, "reason": "ai_disabled"}

    tool_round = 0
    ai_responded = False

    while tool_round < MAX_TOOL_ROUNDS:
        tool_round += 1
        try:
            ai_response = await _call_chat_api(cfg, api_messages)
            record_success()
        except Exception as exc:
            logger.error("AI API 调用失败: room=%s err=%s", room_id, exc)
            await record_error(str(exc))
            break

        choice = (ai_response.get("choices") or [{}])[0]
        message = choice.get("message") or {}
        ai_content = message.get("content")
        tool_calls = message.get("tool_calls")

        # 保存 assistant 消息
        _save_message(db, room_id, "assistant", content=ai_content, tool_calls=tool_calls)
        api_messages.append(message)

        if not tool_calls:
            if ai_content and ai_content.strip():
                ai_responded = True
                logger.info("AI 文本回复 (无工具调用): room=%s content=%s",
                            room_id, (ai_content or "")[:200])
            break

        # ---- 执行工具调用 ----
        for tc in tool_calls:
            func_name = tc.get("function", {}).get("name", "")
            func_args_str = tc.get("function", {}).get("arguments", "{}")
            call_id = tc.get("id", "")

            try:
                func_args = json.loads(func_args_str)
            except Exception:
                func_args = {}

            logger.info("AI 工具调用: room=%s tool=%s args=%s", room_id, func_name, func_args_str[:500])

            tool_result = await _execute_tool(
                func_name, func_args,
                db=db, room_id=room_id, sender_id=sender_id,
                sender_name=sender_name, customer=customer,
                instance_id=instance_id,
                suppress_actions=suppress_actions,
            )

            _save_message(db, room_id, "tool", content=tool_result, name=func_name, tool_call_id=call_id)
            api_messages.append({
                "role": "tool",
                "tool_call_id": call_id,
                "content": tool_result,
            })
            ai_responded = True

    # ---- 清理旧消息 ----
    _trim_history(db, room_id)

    return {"ok": True, "ai_responded": ai_responded}


async def _execute_tool(
    name: str,
    args: dict[str, Any],
    *,
    db: Session,
    room_id: str,
    sender_id: str,
    sender_name: str,
    customer: Optional[dict[str, Any]],
    instance_id: str,
    suppress_actions: bool = False,
) -> str:
    """分发并执行工具调用，返回结果字符串"""
    try:
        if suppress_actions and name in {"create_order_review", "create_modify_review", "void_recent_review", "send_group_message", "mark_followup_continue", "create_cancel_unshipped_review"}:
            return json.dumps({
                "ok": False,
                "suppressed": True,
                "reason": "历史补推消息中检测到员工已接手处理，已阻止机器人重复操作",
            }, ensure_ascii=False)
        if name == "create_order_review":
            return _tool_create_order_review(
                db, room_id, sender_id, sender_name, customer, instance_id, args,
            )
        elif name == "create_modify_review":
            return _tool_create_modify_review(
                db, room_id, sender_id, sender_name, customer, instance_id, args,
            )
        elif name == "void_recent_review":
            return _tool_void_recent_review(db, room_id, args)
        elif name == "query_product_catalog":
            return _tool_query_product_catalog(db, args)
        elif name == "query_product_details":
            return _tool_query_product_details(db, args)
        elif name == "send_group_message":
            return await _tool_send_group_message(room_id, sender_id, instance_id, args)
        elif name == "mark_followup_continue":
            return _tool_mark_followup_continue(db, room_id, args)
        elif name == "create_cancel_unshipped_review":
            return _tool_create_cancel_unshipped_review(
                db, room_id, sender_id, sender_name, customer, instance_id, args,
            )
        elif name == "check_existing_reviews":
            return _tool_check_existing_reviews(db, room_id, args)
        else:
            return json.dumps({"error": f"未知工具: {name}"}, ensure_ascii=False)
    except Exception as exc:
        logger.error("工具执行失败: tool=%s err=%s", name, exc)
        return json.dumps({"error": f"工具执行异常: {exc}"}, ensure_ascii=False)
